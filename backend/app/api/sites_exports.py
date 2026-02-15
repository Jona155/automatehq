import traceback
from copy import copy
from datetime import datetime
from io import BytesIO

from flask import g, request, send_file
from openpyxl import load_workbook

from ..auth_utils import token_required
from .sites import sites_bp, repo, logger, export_service, hours_matrix_service
from .utils import api_response


@sites_bp.route('/<uuid:site_id>/summary/export', methods=['GET'])
@token_required
def export_monthly_summary(site_id):
    site = repo.get_by_id(site_id)
    if not site or site.business_id != g.business_id:
        return api_response(status_code=404, message='Site not found', error='Not Found')
    processing_month = request.args.get('processing_month')
    if not processing_month:
        return api_response(status_code=400, message='processing_month is required', error='Bad Request')

    approved_only = request.args.get('approved_only', 'false').lower() == 'true'
    include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'

    try:
        employees, matrix, status_map, month = hours_matrix_service.load_hours_matrix(site_id, g.business_id, processing_month, approved_only, include_inactive)
        output, download_name = export_service.export_summary_csv(site, employees, matrix, status_map, month)
        return send_file(output, mimetype='text/csv', as_attachment=True, download_name=download_name)
    except ValueError as e:
        return api_response(status_code=400, message='Invalid date format. Use YYYY-MM-DD', error=str(e))
    except Exception as e:
        logger.exception(f'Failed to export hours matrix for site {site_id}')
        traceback.print_exc()
        return api_response(status_code=500, message='Failed to export summary', error=str(e))


@sites_bp.route('/summary/export-batch', methods=['GET'])
@token_required
def export_monthly_summary_batch():
    processing_month = request.args.get('processing_month')
    if not processing_month:
        return api_response(status_code=400, message='processing_month is required', error='Bad Request')

    approved_only = request.args.get('approved_only', 'false').lower() == 'true'
    include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'
    include_inactive_sites = request.args.get('include_inactive_sites', 'false').lower() == 'true'

    try:
        month = datetime.strptime(processing_month, '%Y-%m-%d').date()
    except ValueError as e:
        return api_response(status_code=400, message='Invalid date format. Use YYYY-MM-DD', error=str(e))

    sites = repo.get_all_for_business(g.business_id)
    if not include_inactive_sites:
        sites = [site for site in sites if site.is_active]
    sites = sorted(sites, key=lambda s: ((s.site_name or '').strip().lower(), (s.site_code or '').strip().lower(), str(s.id)))

    try:
        workbook = load_workbook(export_service.resolve_summary_template_path())
    except Exception as e:
        logger.exception('Failed to load summary export template')
        return api_response(status_code=500, message='Failed to load export template', error=str(e))

    template_ws = workbook.worksheets[0]
    style_header = copy(template_ws['B1']._style)
    style_body = copy(template_ws['B3']._style)
    style_total = copy(template_ws['A34']._style)

    for ws in workbook.worksheets[1:]:
        workbook.remove(ws)

    used_sheet_names = set()
    for site in sites:
        try:
            employees, matrix, _, _ = hours_matrix_service.load_hours_matrix(site.id, g.business_id, processing_month, approved_only, include_inactive)
        except Exception:
            logger.exception(f'Failed to build summary for site {site.id}')
            continue

        ws = workbook.copy_worksheet(template_ws)
        ws.title = export_service.safe_sheet_name(site.site_name, used_sheet_names)
        export_service.populate_template_core_sheet(ws, employees, matrix, month, style_header=style_header, style_body=style_body, style_total=style_total)

    if len(workbook.worksheets) > 1:
        workbook.remove(template_ws)
    else:
        export_service.populate_template_core_sheet(template_ws, [], {}, month, style_header=style_header, style_body=style_body, style_total=style_total)
        template_ws.title = 'Sites'

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f"monthly_summary_all_sites_{month.strftime('%Y-%m')}.xlsx")


@sites_bp.route('/<uuid:site_id>/salary-template/export', methods=['GET'])
@token_required
def export_salary_template_site(site_id):
    site = repo.get_by_id(site_id)
    if not site or site.business_id != g.business_id:
        return api_response(status_code=404, message='Site not found', error='Not Found')

    processing_month = request.args.get('processing_month')
    if not processing_month:
        return api_response(status_code=400, message='processing_month is required', error='Bad Request')

    include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'
    try:
        month = datetime.strptime(processing_month, '%Y-%m-%d').date()
        employees, matrix, _, _ = hours_matrix_service.load_hours_matrix(site_id, g.business_id, processing_month, approved_only=False, include_inactive=include_inactive)
        workbook = load_workbook(export_service.resolve_salary_template_path(month))
        ws = workbook.worksheets[0]
        export_service.populate_salary_template_sheet(ws, employees, matrix, month)
        ws.title = export_service.safe_sheet_name(site.site_name, set())
    except ValueError as e:
        return api_response(status_code=500, message='Invalid salary template format', error=str(e))
    except Exception as e:
        logger.exception(f'Failed to export salary template for site {site_id}')
        return api_response(status_code=500, message='Failed to export salary template', error=str(e))

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f"salary_template_{export_service.safe_label(site.site_name) or str(site.id)}_{month.strftime('%Y-%m')}.xlsx")


@sites_bp.route('/salary-template/export-batch', methods=['GET'])
@token_required
def export_salary_template_batch():
    processing_month = request.args.get('processing_month')
    if not processing_month:
        return api_response(status_code=400, message='processing_month is required', error='Bad Request')

    include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'
    include_inactive_sites = request.args.get('include_inactive_sites', 'false').lower() == 'true'

    try:
        month = datetime.strptime(processing_month, '%Y-%m-%d').date()
    except ValueError as e:
        return api_response(status_code=400, message='Invalid date format. Use YYYY-MM-DD', error=str(e))

    sites = repo.get_all_for_business(g.business_id)
    if not include_inactive_sites:
        sites = [site for site in sites if site.is_active]
    sites = sorted(sites, key=lambda s: ((s.site_name or '').strip().lower(), (s.site_code or '').strip().lower(), str(s.id)))

    try:
        workbook = load_workbook(export_service.resolve_salary_template_path(month))
    except Exception as e:
        logger.exception('Failed to load salary export template')
        return api_response(status_code=500, message='Failed to load salary template', error=str(e))

    template_ws = workbook.worksheets[0]
    for ws in workbook.worksheets[1:]:
        workbook.remove(ws)

    used_sheet_names = set()
    populated_count = 0
    for site in sites:
        try:
            employees, matrix, _, _ = hours_matrix_service.load_hours_matrix(site.id, g.business_id, processing_month, approved_only=False, include_inactive=include_inactive)
        except Exception:
            logger.exception(f'Failed to build salary matrix for site {site.id}')
            continue

        ws = workbook.copy_worksheet(template_ws)
        ws.title = export_service.safe_sheet_name(site.site_name, used_sheet_names)
        export_service.populate_salary_template_sheet(ws, employees, matrix, month)
        populated_count += 1

    if populated_count > 0 and len(workbook.worksheets) > 1:
        workbook.remove(template_ws)
    else:
        export_service.populate_salary_template_sheet(template_ws, [], {}, month)
        template_ws.title = 'Sites'

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f"salary_template_all_sites_{month.strftime('%Y-%m')}.xlsx")
