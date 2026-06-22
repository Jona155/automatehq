import logging
import re
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple
from uuid import UUID

import pandas as pd
from flask import Blueprint, request, g, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..auth_utils import token_required, role_required
from ..repositories.site_repository import SiteRepository
from ..repositories.user_repository import UserRepository
from ..utils import normalize_phone
from .employee_imports import _normalize_cell
from .sites import _normalize_contractor_phone, EMAIL_REGEX
from .utils import api_response

logger = logging.getLogger(__name__)

site_tariff_import_bp = Blueprint('site_tariff_import', __name__, url_prefix='/api/sites/tariff-import')

site_repo = SiteRepository()
user_repo = UserRepository()

SITE_NAME_ALIASES = ['שם האתר', 'שם אתר', 'site_name', 'אתר']
TARIFF_ALIASES = ['מחיר', 'תעריף', 'תעריף שעתי', 'tariff', 'hourly_tariff', 'price']
PHONE_ALIASES = ['טלפון איש קשר', 'טלפון', 'מספר טלפון', 'phone', 'contact_phone', 'contractor_phone', 'contractor_phone_number']
EMAIL_ALIASES = ['מייל איש קשר', 'אימייל איש קשר', 'מייל', 'אימייל', 'email', 'contact_email', 'contractor_email', 'contractor_emails']
FIELD_MANAGER_NAME_ALIASES = ['מנהל שדה', 'שם מנהל שדה', 'field_manager', 'field_manager_name']
FIELD_MANAGER_PHONE_ALIASES = ['טלפון מנהל שדה', 'מספר מנהל שדה', 'field_manager_phone']

EMAIL_SPLIT_RE = re.compile(r'[;,]')


def _normalize_name(name: str) -> str:
    """Collapse whitespace for fuzzy matching."""
    return re.sub(r'\s+', ' ', name.strip())


def _build_site_lookups(business_id):
    """Fetch all sites and build lookup dicts by exact and normalized name."""
    sites = site_repo.get_all_for_business(business_id)
    by_name: Dict[str, Any] = {}
    by_normalized: Dict[str, Any] = {}
    for s in sites:
        by_name[s.site_name.strip()] = s
        by_normalized[_normalize_name(s.site_name)] = s
    return sites, by_name, by_normalized


def _build_field_manager_lookups(business_id) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """Build lookups for field-manager resolution, scoped to the business.

    Returns (by_phone, by_id):
      - by_phone: normalized phone -> FIELD_MANAGER user. Phone is the deterministic
        key used to attach a manager to a site. Both sides are run through
        normalize_phone so stored format and Excel-typed format align.
      - by_id: str(user.id) -> user (all users), used to render the manager
        currently assigned to a site.
    """
    all_users = user_repo.get_all_for_business(business_id)
    by_id: Dict[str, Any] = {str(u.id): u for u in all_users}
    by_phone: Dict[str, Any] = {}
    for u in all_users:
        if u.role != 'FIELD_MANAGER':
            continue
        key = normalize_phone(u.phone_number or '')
        if key:
            by_phone[key] = u
    return by_phone, by_id


def _match_site(site_name: str, by_name: Dict[str, Any], by_normalized: Dict[str, Any]):
    """Match a site name: exact-after-trim first, then normalized whitespace."""
    matched = by_name.get(site_name.strip())
    if not matched:
        matched = by_normalized.get(_normalize_name(site_name))
    return matched


def _validate_tariff(tariff_raw) -> Tuple[Optional[float], List[str]]:
    """Validate and parse a tariff cell.

    Returns (parsed_value, errors). Empty / missing cell yields (None, [])
    meaning "no change for this field"; only non-empty values are validated.
    """
    if tariff_raw is None or (isinstance(tariff_raw, str) and not tariff_raw.strip()):
        return None, []
    try:
        value = float(tariff_raw)
        if value < 0:
            return None, ['תעריף חייב להיות חיובי']
        return value, []
    except (ValueError, TypeError):
        return None, ['תעריף לא תקין']


def _validate_phone_cell(phone_raw) -> Tuple[Optional[str], List[str]]:
    """Validate and normalize a phone cell.

    Empty / missing cell yields (None, []) — no change. Otherwise normalises
    via the same logic the per-site UI uses, so manual edits and bulk imports
    produce identical stored values.

    Excel-specific recovery: numeric cells lose their leading zero, so a value
    like 0508123456 arrives here as '508123456'. When the input is purely
    digits, exactly 9 characters long, and starts with 5 (IL mobile prefix
    shape), we restore the leading 0 before handing off to the shared
    normaliser. Formatted input (with separators or '+') is trusted as-is.
    """
    if phone_raw is None:
        return None, []
    if not isinstance(phone_raw, str):
        phone_raw = str(phone_raw)
    raw = phone_raw.strip()
    if not raw:
        return None, []

    if raw.isdigit() and len(raw) == 9 and raw[0] == '5':
        raw = '0' + raw

    digits, err = _normalize_contractor_phone(raw)
    if err:
        return None, [f'מספר טלפון לא תקין: {phone_raw}']
    return digits, []


def _validate_emails_cell(emails_raw) -> Tuple[Optional[List[str]], List[str]]:
    """Validate and normalise an emails cell (semicolon or comma separated)."""
    if emails_raw is None:
        return None, []
    if not isinstance(emails_raw, str):
        emails_raw = str(emails_raw)
    if not emails_raw.strip():
        return None, []

    parts = [p.strip() for p in EMAIL_SPLIT_RE.split(emails_raw)]
    cleaned: List[str] = []
    errors: List[str] = []
    for part in parts:
        if not part:
            continue
        lowered = part.lower()
        if not EMAIL_REGEX.match(lowered):
            errors.append(f'כתובת מייל לא תקינה: {part}')
            continue
        if lowered not in cleaned:
            cleaned.append(lowered)
    if errors:
        return None, errors
    return cleaned, []


def _tariff_changed(current: Optional[float], new: Optional[float]) -> bool:
    """Compare tariffs safely, handling Decimal->float rounding.

    `new is None` means "leave the field as-is" — never reported as a change.
    """
    if new is None:
        return False
    if current is None:
        return True
    return round(current, 2) != round(new, 2)


def _phone_changed(current: Optional[str], new: Optional[str]) -> bool:
    if new is None:
        return False
    return (current or '') != (new or '')


def _emails_changed(current: Optional[List[str]], new: Optional[List[str]]) -> bool:
    if new is None:
        return False
    cur = [e.lower() for e in (current or [])]
    return sorted(cur) != sorted(new)


def _field_manager_changed(current_id: Optional[str], new_id: Optional[str], present: bool) -> bool:
    """Whether the field-manager assignment changes.

    Only meaningful when the field-manager-phone column is present. When absent,
    the field is left untouched and never reported as a change.
    """
    if not present:
        return False
    return (current_id or None) != (new_id or None)


def _summarize(rows: List[Dict[str, Any]]) -> Dict[str, int]:
    summary = {'update': 0, 'no_change': 0, 'error': 0, 'total': len(rows)}
    for row in rows:
        action = row.get('action')
        if action in summary:
            summary[action] += 1
        else:
            summary['error'] += 1
    return summary


def _find_column(row_values: List[str], aliases: List[str]) -> Optional[int]:
    for ci, cell in enumerate(row_values):
        if cell in aliases:
            return ci
    return None


def _parse_tariff_file(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Parse Excel file, auto-detecting header row and (optional) columns.

    Required: a header row with the site-name column. At least one of
    tariff/phone/email columns must be present (otherwise the file has nothing
    to import). Empty cells in a present column are treated as "no change" by
    the diff stage rather than as errors — this lets users round-trip the
    export and edit only the fields they care about.
    """
    try:
        df = pd.read_excel(BytesIO(file_bytes), header=None, sheet_name='תעריפים לכל אתר')
    except (KeyError, ValueError):
        df = pd.read_excel(BytesIO(file_bytes), header=None, sheet_name=0)

    header_row = None
    site_col = None
    tariff_col = None
    phone_col = None
    email_col = None
    fm_name_col = None
    fm_phone_col = None

    for idx, row in df.iterrows():
        row_values = [str(v).strip() if pd.notna(v) else '' for v in row]
        site_candidate = _find_column(row_values, SITE_NAME_ALIASES)
        if site_candidate is None:
            continue
        tariff_candidate = _find_column(row_values, TARIFF_ALIASES)
        phone_candidate = _find_column(row_values, PHONE_ALIASES)
        email_candidate = _find_column(row_values, EMAIL_ALIASES)
        fm_name_candidate = _find_column(row_values, FIELD_MANAGER_NAME_ALIASES)
        fm_phone_candidate = _find_column(row_values, FIELD_MANAGER_PHONE_ALIASES)
        if (
            tariff_candidate is None and phone_candidate is None
            and email_candidate is None and fm_phone_candidate is None
        ):
            continue
        header_row = idx
        site_col = site_candidate
        tariff_col = tariff_candidate
        phone_col = phone_candidate
        email_col = email_candidate
        fm_name_col = fm_name_candidate
        fm_phone_col = fm_phone_candidate
        break

    if header_row is None:
        raise ValueError('לא נמצאו כותרות מתאימות בקובץ (שם האתר ולפחות אחת מ: תעריף / טלפון / מייל / טלפון מנהל שדה)')

    # Whether the field-manager-phone column is present at all in the file.
    # This gates the "empty cell = unassign" behaviour: a file without the
    # column must never touch existing field-manager assignments.
    field_manager_present = fm_phone_col is not None

    rows = []
    for idx in range(header_row + 1, len(df)):
        site_name = _normalize_cell(df.iloc[idx, site_col])
        if not site_name:
            continue

        rows.append({
            'row_number': idx + 1,
            'site_name': site_name,
            'tariff_raw': _normalize_cell(df.iloc[idx, tariff_col]) if tariff_col is not None else None,
            'phone_raw': _normalize_cell(df.iloc[idx, phone_col]) if phone_col is not None else None,
            'emails_raw': _normalize_cell(df.iloc[idx, email_col]) if email_col is not None else None,
            'field_manager_name_raw': _normalize_cell(df.iloc[idx, fm_name_col]) if fm_name_col is not None else None,
            'field_manager_phone_raw': _normalize_cell(df.iloc[idx, fm_phone_col]) if fm_phone_col is not None else None,
            'field_manager_present': field_manager_present,
        })

    return rows


def _resolve_field_manager(
    matched_site: Any,
    fm_present: bool,
    fm_phone_raw: Optional[str],
    fm_name_raw: Optional[str],
    fm_by_phone: Dict[str, Any],
    fm_by_id: Dict[str, Any],
    errors: List[str],
    warnings: List[str],
) -> Dict[str, Any]:
    """Resolve the field-manager target for a row by phone.

    Phone is the deterministic key. An empty cell (when the column is present)
    means "unassign". A non-empty phone that resolves to no FIELD_MANAGER in the
    business is a blocking error (`unresolved=True`) — the caller fails the whole
    import. The name in the file is informational: a mismatch is a warning only.
    """
    # field_manager_phone_input echoes the phone exactly as entered in the file so
    # the value round-trips to /apply, which re-resolves it server-side (a phone
    # that resolved nowhere must still fail on apply, not be mistaken for unassign).
    result = {
        'current_field_manager_id': None,
        'current_field_manager_name': None,
        'current_field_manager_phone': None,
        'new_field_manager_id': None,
        'new_field_manager_name': None,
        'new_field_manager_phone': None,
        'field_manager_phone_input': (fm_phone_raw or '').strip() if fm_present else None,
        'field_manager_present': fm_present,
        'field_manager_unresolved': False,
    }

    if matched_site is not None and matched_site.field_manager_id:
        current_mgr = fm_by_id.get(str(matched_site.field_manager_id))
        if current_mgr is not None:
            result['current_field_manager_id'] = str(current_mgr.id)
            result['current_field_manager_name'] = current_mgr.full_name
            result['current_field_manager_phone'] = current_mgr.phone_number or None

    if not fm_present:
        return result

    raw = (fm_phone_raw or '').strip()
    if not raw:
        # Column present, cell empty -> unassign (new target stays None).
        return result

    mgr = fm_by_phone.get(normalize_phone(raw))
    if mgr is None:
        result['field_manager_unresolved'] = True
        errors.append(f'מנהל שדה לא נמצא עבור טלפון: {raw}')
        return result

    result['new_field_manager_id'] = str(mgr.id)
    result['new_field_manager_name'] = mgr.full_name
    result['new_field_manager_phone'] = mgr.phone_number or None

    file_name = (fm_name_raw or '').strip()
    if file_name and _normalize_name(file_name) != _normalize_name(mgr.full_name or ''):
        warnings.append(f'שם מנהל השדה בקובץ ("{file_name}") שונה מהשם הרשום ("{mgr.full_name}")')

    return result


def _build_diff_row(
    site_name: str,
    row_number: int,
    new_tariff: Optional[float],
    new_phone: Optional[str],
    new_emails: Optional[List[str]],
    field_errors: List[str],
    by_name: Dict,
    by_normalized: Dict,
    fm_present: bool = False,
    fm_phone_raw: Optional[str] = None,
    fm_name_raw: Optional[str] = None,
    fm_by_phone: Optional[Dict[str, Any]] = None,
    fm_by_id: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    errors = list(field_errors)
    warnings: List[str] = []

    matched_site = _match_site(site_name, by_name, by_normalized)

    matched_site_id = None
    matched_site_name = None
    current_tariff = None
    current_phone = None
    current_emails = None
    action = 'error'

    fm = _resolve_field_manager(
        matched_site, fm_present, fm_phone_raw, fm_name_raw,
        fm_by_phone or {}, fm_by_id or {}, errors, warnings,
    )

    if not matched_site:
        errors.append('אתר לא נמצא')
    else:
        matched_site_id = str(matched_site.id)
        matched_site_name = matched_site.site_name
        current_tariff = float(matched_site.hourly_tariff) if matched_site.hourly_tariff is not None else None
        current_phone = matched_site.contractor_phone_number or None
        current_emails = list(matched_site.contractor_emails or [])

        if not errors:
            any_change = (
                _tariff_changed(current_tariff, new_tariff)
                or _phone_changed(current_phone, new_phone)
                or _emails_changed(current_emails, new_emails)
                or _field_manager_changed(
                    fm['current_field_manager_id'], fm['new_field_manager_id'], fm_present
                )
            )
            action = 'update' if any_change else 'no_change'

    if errors:
        action = 'error'

    row = {
        'row_number': row_number,
        'site_name_from_file': site_name,
        'matched_site_id': matched_site_id,
        'matched_site_name': matched_site_name,
        'current_tariff': current_tariff,
        'new_tariff': new_tariff,
        'current_phone': current_phone,
        'new_phone': new_phone,
        'current_emails': current_emails,
        'new_emails': new_emails,
        'action': action,
        'errors': errors,
        'warnings': warnings,
        '_matched_site': matched_site,  # internal, stripped before response
    }
    row.update(fm)
    return row


def _build_tariff_diff(parsed_rows: List[Dict[str, Any]], business_id) -> Tuple[List[Dict[str, Any]], Dict[str, int], bool]:
    _sites, by_name, by_normalized = _build_site_lookups(business_id)
    fm_by_phone, fm_by_id = _build_field_manager_lookups(business_id)

    seen_names: Dict[str, int] = {}
    diff_rows: List[Dict[str, Any]] = []

    for parsed in parsed_rows:
        site_name = parsed['site_name']
        new_tariff, tariff_errors = _validate_tariff(parsed.get('tariff_raw'))
        new_phone, phone_errors = _validate_phone_cell(parsed.get('phone_raw'))
        new_emails, email_errors = _validate_emails_cell(parsed.get('emails_raw'))
        field_errors = tariff_errors + phone_errors + email_errors

        row_entry = _build_diff_row(
            site_name, parsed['row_number'],
            new_tariff, new_phone, new_emails,
            field_errors, by_name, by_normalized,
            fm_present=parsed.get('field_manager_present', False),
            fm_phone_raw=parsed.get('field_manager_phone_raw'),
            fm_name_raw=parsed.get('field_manager_name_raw'),
            fm_by_phone=fm_by_phone,
            fm_by_id=fm_by_id,
        )

        norm_key = _normalize_name(site_name)
        if norm_key in seen_names:
            prev_idx = seen_names[norm_key]
            diff_rows[prev_idx]['action'] = 'error'
            diff_rows[prev_idx]['errors'].append('שורה כפולה - נדרש רק רשומה אחת לכל אתר')

        seen_names[norm_key] = len(diff_rows)
        diff_rows.append(row_entry)

    # Whole-import-fail: any unresolved field-manager phone blocks the entire import.
    blocked = any(row.get('field_manager_unresolved') for row in diff_rows)

    for row in diff_rows:
        row.pop('_matched_site', None)

    return diff_rows, _summarize(diff_rows), blocked


@site_tariff_import_bp.route('/preview', methods=['POST'])
@token_required
@role_required('ADMIN')
def preview_tariff_import():
    if 'file' not in request.files:
        return api_response(status_code=400, message="No file provided", error="Bad Request")

    file = request.files['file']
    if not file or file.filename == '':
        return api_response(status_code=400, message="No file selected", error="Bad Request")

    try:
        parsed_rows = _parse_tariff_file(file.read())
        diff_rows, summary, blocked = _build_tariff_diff(parsed_rows, g.business_id)

        block_reason = 'ייבוא נחסם — מנהל שדה לא זוהה לפי מספר טלפון באחת או יותר מהשורות' if blocked else None
        return api_response(data={'summary': summary, 'rows': diff_rows, 'blocked': blocked, 'block_reason': block_reason})
    except ValueError as e:
        return api_response(status_code=400, message=str(e), error="Bad Request")
    except Exception as e:
        logger.exception("Failed to preview site tariff import")
        return api_response(status_code=500, message="Failed to preview site tariff import", error=str(e))


@site_tariff_import_bp.route('/apply', methods=['POST'])
@token_required
@role_required('ADMIN')
def apply_tariff_import():
    data = request.get_json()
    if not data or 'rows' not in data:
        return api_response(status_code=400, message="rows is required", error="Bad Request")

    rows_input = data['rows']
    if not isinstance(rows_input, list):
        return api_response(status_code=400, message="rows must be an array", error="Bad Request")

    try:
        _sites, by_name, by_normalized = _build_site_lookups(g.business_id)
        fm_by_phone, fm_by_id = _build_field_manager_lookups(g.business_id)

        # Pass 1: validate & build all rows without mutating anything. Each entry
        # keeps a handle on the matched site plus the resolved new field values.
        diff_rows = []
        plan = []  # parallel list of (row_entry, matched_site, new_tariff, new_phone, new_emails)
        seen_names: Dict[str, int] = {}

        for row in rows_input:
            site_name = row.get('site_name_from_file', '')
            row_number = row.get('row_number')

            if not site_name:
                continue

            new_tariff, tariff_errors = _validate_tariff(row.get('new_tariff'))
            new_phone, phone_errors = _validate_phone_cell(row.get('new_phone'))
            # new_emails arrives as either a list (from frontend) or a string (defensive)
            raw_emails = row.get('new_emails')
            if isinstance(raw_emails, list):
                raw_emails = '; '.join(str(e) for e in raw_emails)
            new_emails, email_errors = _validate_emails_cell(raw_emails)
            field_errors = tariff_errors + phone_errors + email_errors

            row_entry = _build_diff_row(
                site_name, row_number,
                new_tariff, new_phone, new_emails,
                field_errors, by_name, by_normalized,
                fm_present=bool(row.get('field_manager_present')),
                fm_phone_raw=row.get('field_manager_phone_input'),
                fm_by_phone=fm_by_phone,
                fm_by_id=fm_by_id,
            )

            norm_key = _normalize_name(site_name)
            if norm_key in seen_names:
                prev_idx = seen_names[norm_key]
                diff_rows[prev_idx]['action'] = 'error'
                diff_rows[prev_idx]['errors'].append('שורה כפולה - נדרש רק רשומה אחת לכל אתר')

            seen_names[norm_key] = len(diff_rows)

            matched_site = row_entry.pop('_matched_site', None)
            diff_rows.append(row_entry)
            plan.append((row_entry, matched_site, new_tariff, new_phone, new_emails))

        # Whole-import-fail: if any field-manager phone is unresolved, apply nothing.
        if any(row.get('field_manager_unresolved') for row in diff_rows):
            site_repo.rollback()
            return api_response(
                status_code=400,
                message='ייבוא נחסם — מנהל שדה לא זוהה לפי מספר טלפון באחת או יותר מהשורות. לא בוצעו שינויים.',
                error="Bad Request",
            )

        # Pass 2: apply mutations for update rows, then commit once.
        applied = []
        for row_entry, matched_site, new_tariff, new_phone, new_emails in plan:
            if row_entry['action'] != 'update' or not matched_site:
                continue

            site_changes: Dict[str, Any] = {}
            if _tariff_changed(row_entry['current_tariff'], new_tariff):
                matched_site.hourly_tariff = new_tariff
                site_changes['tariff'] = {'old': row_entry['current_tariff'], 'new': new_tariff}
            if _phone_changed(row_entry['current_phone'], new_phone):
                matched_site.contractor_phone_number = new_phone
                site_changes['phone'] = {'old': row_entry['current_phone'], 'new': new_phone}
            if _emails_changed(row_entry['current_emails'], new_emails):
                matched_site.contractor_emails = new_emails
                site_changes['emails'] = {'old': row_entry['current_emails'], 'new': new_emails}
            if _field_manager_changed(
                row_entry['current_field_manager_id'],
                row_entry['new_field_manager_id'],
                row_entry['field_manager_present'],
            ):
                new_fm_id = row_entry['new_field_manager_id']
                matched_site.field_manager_id = UUID(new_fm_id) if new_fm_id else None
                site_changes['field_manager'] = {
                    'old': row_entry['current_field_manager_name'],
                    'new': row_entry['new_field_manager_name'],
                }

            if site_changes:
                applied.append({
                    'site_id': str(matched_site.id),
                    'site_name': matched_site.site_name,
                    'changes': site_changes,
                })

        site_repo.commit()

        return api_response(
            data={'summary': _summarize(diff_rows), 'rows': diff_rows, 'applied': applied},
            message="Site tariff import applied"
        )
    except Exception as e:
        site_repo.rollback()
        logger.exception("Failed to apply site tariff import")
        return api_response(status_code=500, message="Failed to apply site tariff import", error=str(e))


@site_tariff_import_bp.route('/export', methods=['GET'])
@token_required
@role_required('ADMIN')
def export_tariffs():
    """Export site details as an Excel file matching the import format.

    Round-trips through /preview + /apply: download → edit → re-upload.
    Includes site name, hourly tariff, contact phone, contact emails and the
    assigned field manager (name + phone).
    """
    try:
        include_inactive = request.args.get('include_inactive', 'true').lower() == 'true'
        sites = site_repo.get_all_for_business(g.business_id)
        if not include_inactive:
            sites = [s for s in sites if s.is_active]

        sites = sorted(sites, key=lambda s: (s.site_name or '').strip().lower())

        # Resolve assigned field managers (the phone is the import key).
        users_by_id = {str(u.id): u for u in user_repo.get_all_for_business(g.business_id)}

        workbook = Workbook()
        ws = workbook.active
        ws.title = 'תעריפים לכל אתר'
        ws.sheet_view.rightToLeft = True

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='1F4E78')
        center = Alignment(horizontal='center', vertical='center')

        headers = ['שם האתר', 'תעריף שעתי', 'טלפון איש קשר', 'מייל איש קשר', 'מנהל שדה', 'טלפון מנהל שדה']
        for col_idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for row_idx, site in enumerate(sites, start=2):
            ws.cell(row=row_idx, column=1, value=site.site_name)

            tariff_cell = ws.cell(row=row_idx, column=2)
            if site.hourly_tariff is not None:
                tariff_cell.value = float(site.hourly_tariff)
                tariff_cell.number_format = '0.00'

            phone_cell = ws.cell(row=row_idx, column=3)
            # Force text format so Excel won't strip a leading 0 from local-format
            # numbers users might type when editing (e.g. 0508... → 508...).
            phone_cell.number_format = '@'
            if site.contractor_phone_number:
                phone_cell.value = f'+{site.contractor_phone_number}'

            emails = site.contractor_emails or []
            if emails:
                ws.cell(row=row_idx, column=4, value='; '.join(emails))

            manager = users_by_id.get(str(site.field_manager_id)) if site.field_manager_id else None
            if manager is not None:
                ws.cell(row=row_idx, column=5, value=manager.full_name)
            fm_phone_cell = ws.cell(row=row_idx, column=6)
            # Text format so the manager's local-format phone keeps its leading 0.
            fm_phone_cell.number_format = '@'
            if manager is not None and manager.phone_number:
                fm_phone_cell.value = manager.phone_number

        ws.column_dimensions[get_column_letter(1)].width = 36
        ws.column_dimensions[get_column_letter(2)].width = 14
        ws.column_dimensions[get_column_letter(3)].width = 20
        ws.column_dimensions[get_column_letter(4)].width = 40
        ws.column_dimensions[get_column_letter(5)].width = 24
        ws.column_dimensions[get_column_letter(6)].width = 20
        ws.freeze_panes = 'A2'

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='site_details.xlsx',
        )
    except Exception as e:
        logger.exception("Failed to export site details")
        return api_response(status_code=500, message="Failed to export site details", error=str(e))
