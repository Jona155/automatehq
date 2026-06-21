from flask import Blueprint, request, g, send_file
import os
import uuid
import traceback
import logging
import re
import time
from datetime import datetime, timedelta, timezone
import secrets
import csv
import calendar
from io import BytesIO, StringIO
import unicodedata
from pathlib import Path
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, or_, func, case
from sqlalchemy.orm import joinedload
from twilio.rest import Client
from ..repositories.site_repository import SiteRepository
from ..repositories.employee_repository import EmployeeRepository
from ..repositories.upload_access_request_repository import UploadAccessRequestRepository
from ..repositories.business_repository import BusinessRepository
from ..repositories.user_repository import UserRepository
from ..services.sites.hours_matrix_service import (
    build_employee_upload_status_map,
    get_latest_work_card_with_extraction_by_employee,
)
from .utils import api_response, model_to_dict, models_to_list
from .dashboard import invalidate_business_cache
from ..auth_utils import token_required, role_required
from ..utils import normalize_phone
from ..models.work_cards import WorkCard, WorkCardExtraction, WorkCardDayEntry
from ..models.sites import Employee
from ..extensions import db
from ..observability import QueryCounter, sites_metrics
from ..services.email_service import send_email_with_attachment
from ..services.whatsapp_listener_client import (
    WhatsAppAuthError,
    WhatsAppBadRequestError,
    WhatsAppListenerClient,
    WhatsAppListenerError,
    WhatsAppNotConnectedError,
    WhatsAppNumberNotRegisteredError,
    WhatsAppPayloadTooLargeError,
)

logger = logging.getLogger(__name__)

sites_bp = Blueprint('sites', __name__, url_prefix='/api/sites')
repo = SiteRepository()
employee_repo = EmployeeRepository()
access_repo = UploadAccessRequestRepository()
business_repo = BusinessRepository()
user_repo = UserRepository()

STATUS_LABELS = {
    'APPROVED': 'מאושר',
    'NEEDS_REVIEW': 'ממתין לסקירה',
    'NEEDS_ASSIGNMENT': 'ממתין לשיוך',
    'REJECTED': 'נדחה',
    'NO_UPLOAD': 'ללא העלאה',
}

STATUS_DAY_LABELS = {
    'VACATION': 'חופשה',
    'SICK': 'מחלה',
    'INTERNATIONAL_VISA': 'ויזה בינלאומית',
    'HOLIDAY': 'חג',
}

EMAIL_REGEX = re.compile(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$')

IL_COUNTRY_CODE = '972'


def _normalize_contractor_phone(raw):
    """Normalize a contractor phone input to E.164 digits (no leading +).

    Accepts user input like '050-123-4567', '+972 50 123 4567', '0501234567'.
    Assumes Israel (+972) when input is a local number starting with '0'.
    Returns (digits_only, error_msg). If raw is None/empty, returns (None, None).
    """
    if raw is None:
        return None, None
    if not isinstance(raw, str):
        return None, 'contractor_phone_number must be a string'

    trimmed = raw.strip()
    if not trimmed:
        return None, None

    has_plus = trimmed.startswith('+')
    has_double_zero = trimmed.startswith('00')
    digits = re.sub(r'\D', '', trimmed)
    if not digits:
        return None, f'Invalid phone number: {raw}'

    if has_plus:
        pass
    elif has_double_zero:
        digits = digits[2:]
    elif digits.startswith('0'):
        digits = IL_COUNTRY_CODE + digits[1:]

    if not (8 <= len(digits) <= 15):
        return None, f'Invalid phone number: {raw}'

    return digits, None


def _validate_contractor_emails(data):
    """Validate and normalize contractor_emails field. Returns (cleaned_list, error_msg)."""
    raw = data.get('contractor_emails')
    if raw is None:
        return None, None
    if not isinstance(raw, list):
        return None, 'contractor_emails must be a list'
    cleaned = []
    for item in raw:
        if not isinstance(item, str):
            return None, 'Each contractor email must be a string'
        email = item.strip().lower()
        if not email:
            continue
        if not EMAIL_REGEX.match(email):
            return None, f'Invalid email address: {item}'
        cleaned.append(email)
    return cleaned, None

def _generate_access_token():
    token = None
    for _ in range(5):
        candidate = secrets.token_urlsafe(32)
        if not access_repo.token_exists(candidate):
            token = candidate
            break
    return token

def _format_whatsapp_number(raw_phone: str):
    if not raw_phone:
        return None
    if raw_phone.startswith('0'):
        return '+972' + raw_phone[1:]
    return '+' + raw_phone

def _build_access_link_url(token: str):
    return f"{request.host_url.rstrip('/')}/portal/{token}"

def _safe_label(value: str) -> str:
    if not value:
        return ''
    normalized = unicodedata.normalize('NFKC', value)
    cleaned = []
    for ch in normalized:
        cat = unicodedata.category(ch)
        if cat[0] in {'L', 'N'}:
            cleaned.append(ch)
        else:
            cleaned.append('_')
    label = ''.join(cleaned).strip('_')
    label = '_'.join(filter(None, label.split('_')))
    return label

def _safe_sheet_name(value: str, existing: set) -> str:
    if not value:
        base = 'Site'
    else:
        base = value
    invalid = ['\\', '/', '*', '[', ']', ':', '?']
    for ch in invalid:
        base = base.replace(ch, ' ')
    base = ' '.join(base.split()).strip()
    if not base:
        base = 'Site'
    # Trailing spaces are stripped by Excel when a workbook is saved, so a sheet
    # title that ends mid-word at the 31-char limit must be rstripped here to
    # stay round-trippable on re-import.
    base = base[:31].rstrip()
    candidate = base
    counter = 2
    while candidate in existing:
        suffix = f" {counter}"
        candidate = (base[:31 - len(suffix)] + suffix).rstrip()
        counter += 1
    existing.add(candidate)
    return candidate

def _resolve_summary_template_path() -> Path:
    """Resolve the Excel template used for batch site export."""
    base_dir = Path(__file__).resolve().parents[2]
    templates_dir = base_dir / 'excel_extraction_example'
    preferred = templates_dir / 'שעות עבודה לפי אתרים ינואר 26 (1).xlsx'
    if preferred.exists():
        return preferred

    candidates = sorted(templates_dir.glob('*.xlsx'))
    if not candidates:
        raise FileNotFoundError('No Excel template found in backend/excel_extraction_example')
    return candidates[0]

def _resolve_salary_template_path(month_date) -> Path:
    """Resolve salary export template for a month from employee_sheet_extraction."""
    root_dir = Path(__file__).resolve().parents[3]
    templates_dir = root_dir / 'employee_sheet_extraction'
    if not templates_dir.exists():
        raise FileNotFoundError('Template folder employee_sheet_extraction was not found')

    month_token = f"{month_date.month:02d}_{month_date.year}"
    month_candidates = []
    generic_candidates = []

    for candidate in sorted(templates_dir.glob('*.xlsx')):
        stem = candidate.stem.lower()
        if 'worker_new_hours_template' not in stem:
            continue
        generic_candidates.append(candidate)
        if month_token in stem or f"{month_date.month}_{month_date.year}" in stem:
            month_candidates.append(candidate)

    if month_candidates:
        return month_candidates[-1]
    if generic_candidates:
        return generic_candidates[-1]

    raise FileNotFoundError(
        f'No salary template found in {templates_dir} for {month_date.strftime("%Y-%m")}'
    )


def _extract_salary_day_columns_map(ws, month_date):
    """
    Parse row-1 day headers in salary template and return {day_number: column_index}.
    Expected header format per day column: "D.MM", e.g. "1.02".
    """
    day_columns = {}
    observed_months = set()
    pattern = re.compile(r'^\s*(\d{1,2})\.(\d{1,2})\s*$')

    for col in range(2, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value is None:
            continue
        match = pattern.match(str(value))
        if not match:
            continue
        day = int(match.group(1))
        month = int(match.group(2))
        if day < 1 or day > 31:
            continue
        observed_months.add(month)
        day_columns[day] = col

    if not day_columns:
        raise ValueError('Invalid salary template format: missing day columns on header row')
    return day_columns


def _copy_salary_column_template(ws, source_col, target_col):
    """Copy style and width from one column to another across the worksheet."""
    source_letter = get_column_letter(source_col)
    target_letter = get_column_letter(target_col)
    ws.column_dimensions[target_letter].width = ws.column_dimensions[source_letter].width

    for row in range(1, ws.max_row + 1):
        source_cell = ws.cell(row=row, column=source_col)
        target_cell = ws.cell(row=row, column=target_col)
        target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)
        target_cell.value = source_cell.value


def _ensure_salary_template_month_columns(ws, month_date):
    """
    Ensure the salary sheet day columns match the requested month:
    - adjust number of day columns (insert/delete at the end of the day range)
    - rewrite header values to D.MM for requested month
    """
    parsed = _extract_salary_day_columns_map(ws, month_date)
    day_start_col = min(parsed.values())
    existing_days = max(parsed.keys())
    target_days = calendar.monthrange(month_date.year, month_date.month)[1]

    if existing_days < target_days:
        cols_to_add = target_days - existing_days
        insert_at = day_start_col + existing_days
        ws.insert_cols(insert_at, cols_to_add)
        for offset in range(cols_to_add):
            col_index = insert_at + offset
            _copy_salary_column_template(ws, insert_at - 1, col_index)
    elif existing_days > target_days:
        delete_start = day_start_col + target_days
        delete_count = existing_days - target_days
        ws.delete_cols(delete_start, delete_count)

    day_columns = {}
    for day in range(1, target_days + 1):
        col = day_start_col + day - 1
        ws.cell(row=1, column=col, value=f'{day}.{month_date.month:02d}')
        day_columns[day] = col
    return day_columns


def _find_salary_instruction_row(ws):
    """Find first row that starts the instructions block (contains 'הוראות' in column A)."""
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value is None:
            continue
        if 'הוראות' in str(cell_value):
            return row
    raise ValueError('Invalid salary template format: could not find instructions row')


def _copy_salary_row_template(ws, source_row, target_row):
    """Copy style and baseline values from one template row to another."""
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)
        target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)
        target_cell.value = source_cell.value

    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _populate_salary_template_sheet(ws, employees, matrix, month_date, status_matrix=None):
    """
    Populate salary template:
    - Row 1 remains as header (passport/day columns).
    - Employee rows start at row 2 and continue until instructions row.
    - Column A contains employee passport/ID.
    - Day columns contain numeric hours when available, otherwise remain blank
      (except prefilled 'שבת' values are preserved).
    """
    if status_matrix is None:
        status_matrix = {}
    ws.sheet_view.rightToLeft = True
    day_columns = _ensure_salary_template_month_columns(ws, month_date)
    days_in_month = calendar.monthrange(month_date.year, month_date.month)[1]
    employee_start_row = 2
    instruction_row = _find_salary_instruction_row(ws)
    base_template_row = max(employee_start_row, instruction_row - 1)
    available_rows = max(0, instruction_row - employee_start_row)
    needed_rows = len(employees)

    if needed_rows > available_rows:
        rows_to_insert = needed_rows - available_rows
        ws.insert_rows(instruction_row, rows_to_insert)
        for row in range(instruction_row, instruction_row + rows_to_insert):
            _copy_salary_row_template(ws, base_template_row, row)
        instruction_row += rows_to_insert

    def clear_row(row_index):
        ws.cell(row=row_index, column=1, value=None)
        for day, col in day_columns.items():
            cell = ws.cell(row=row_index, column=col)
            is_saturday = day <= days_in_month and datetime(month_date.year, month_date.month, day).weekday() == 5
            cell.value = 'שבת' if is_saturday else None

    for idx, employee in enumerate(employees):
        row_index = employee_start_row + idx
        employee_id_value = (employee.passport_id or '').strip() if employee.passport_id else ''
        ws.cell(row=row_index, column=1, value=employee_id_value)

        employee_id_str = str(employee.id)
        employee_days = matrix.get(employee_id_str, {})
        employee_statuses = status_matrix.get(employee_id_str, {})
        for day, col in day_columns.items():
            cell = ws.cell(row=row_index, column=col)
            status = employee_statuses.get(day)
            if status:
                cell.value = STATUS_DAY_LABELS[status]
            else:
                hours = employee_days.get(day)
                if hours is None:
                    is_saturday = day <= days_in_month and datetime(month_date.year, month_date.month, day).weekday() == 5
                    cell.value = 'שבת' if is_saturday else None
                else:
                    cell.value = round(float(hours), 2)

    for row_index in range(employee_start_row + needed_rows, instruction_row):
        clear_row(row_index)


def _sort_employees_for_export(employees):
    return sorted(
        employees,
        key=lambda e: (
            (e.full_name or '').strip().lower(),
            (e.passport_id or '').strip().lower(),
            str(e.id)
        )
    )


def _format_day_label(year: int, month: int, day: int, days_in_month: int):
    if day <= days_in_month and datetime(year, month, day).weekday() == 5:
        return f'{day}-שבת'
    return day


def _day_fallback_value(year: int, month: int, day: int, days_in_month: int):
    if day > days_in_month:
        return None
    if datetime(year, month, day).weekday() == 5:
        return 'שבת'
    return None


def _populate_template_core_sheet(
    ws,
    employees,
    matrix,
    month_date,
    style_header,
    style_body,
    style_total,
    status_matrix=None,
    monthly_totals=None,
):
    """Populate a worksheet in the core template format (no fee summary rows).

    When monthly_totals[employee_id] is set, the row-34 total cell for that
    employee is written as a literal value (the manual override) instead of
    the default =SUM(C3:C33) formula. Per-day cells still reflect any
    extracted/manual day entries — only the total cell is overridden.
    """
    if status_matrix is None:
        status_matrix = {}
    if monthly_totals is None:
        monthly_totals = {}
    days_in_month = calendar.monthrange(month_date.year, month_date.month)[1]
    employee_count = len(employees)
    last_data_col = max(2, employee_count + 1)

    # Clear template values (including fee/footer cells) while preserving worksheet settings.
    clear_max_col = max(last_data_col, ws.max_column)
    for row in range(1, 39):
        for col in range(1, clear_max_col + 1):
            ws.cell(row=row, column=col).value = None

    ws.sheet_view.rightToLeft = True

    ws.cell(row=1, column=1, value='יום בחודש')._style = copy(style_header)
    ws.cell(row=2, column=1, value=None)._style = copy(style_body)

    for idx, employee in enumerate(employees, start=2):
        passport_value = (employee.passport_id or '').strip() if employee.passport_id else ''
        ws.cell(row=1, column=idx, value=passport_value)._style = copy(style_header)
        first_name = (employee.full_name or '').split()[0] if employee.full_name else ''
        ws.cell(row=2, column=idx, value=first_name)._style = copy(style_body)

    # Body rows: one row per actual day in the month (rows 3..days_in_month+2)
    for day in range(1, days_in_month + 1):
        row = day + 2
        ws.cell(
            row=row,
            column=1,
            value=_format_day_label(month_date.year, month_date.month, day, days_in_month)
        )._style = copy(style_body)

        for idx, employee in enumerate(employees, start=2):
            employee_id_str = str(employee.id)
            status = status_matrix.get(employee_id_str, {}).get(day)
            if status:
                value = STATUS_DAY_LABELS[status]
            else:
                employee_days = matrix.get(employee_id_str, {})
                value = employee_days.get(day)
                if value is None:
                    value = _day_fallback_value(month_date.year, month_date.month, day, days_in_month)
            ws.cell(row=row, column=idx, value=value)._style = copy(style_body)

    # Clear template styling for any excess rows (e.g. days 29-31 when month has 28 days)
    for excess_day in range(days_in_month + 1, 32):
        excess_row = excess_day + 2
        for col in range(1, clear_max_col + 1):
            cell = ws.cell(row=excess_row, column=col)
            cell.value = None
            cell._style = copy(style_body)

    total_row = days_in_month + 3
    last_day_row = days_in_month + 2
    ws.cell(row=total_row, column=1, value='סה"כ')._style = copy(style_total)
    for col in range(2, last_data_col + 1):
        col_letter = get_column_letter(col)
        employee_idx = col - 2
        employee = employees[employee_idx] if employee_idx < len(employees) else None
        override = monthly_totals.get(str(employee.id)) if employee is not None else None
        if override is not None:
            cell_value = round(float(override), 2)
        else:
            cell_value = f'=SUM({col_letter}3:{col_letter}{last_day_row})'
        ws.cell(row=total_row, column=col, value=cell_value)._style = copy(style_total)

    return total_row

def _add_tariff_summary(ws, employee_count, hourly_tariff, style_body,
                        style_site_total, style_tariff, style_tariff_label,
                        total_row=34):
    """Add tariff/fee summary rows to a populated site sheet.

    Clears the template's fixed colored cells and, when the site has an
    hourly tariff, writes grand-total, price-per-hour, cost-without-VAT
    and cost-with-VAT at dynamic columns based on employee count.

    total_row: the row written by _populate_template_core_sheet for the
    'סה"כ' totals; defaults to 34 (31-day month). For shorter months this
    shifts up so that tariff rows follow immediately after the total.
    """
    TMPL_VALUE_COL = 8
    TMPL_LABEL_COL = 9

    last_data_col = max(2, employee_count + 1)
    value_col = last_data_col + 1
    label_col = value_col + 1

    label_row = total_row - 1   # last day row — used for site total label
    tariff_row = total_row + 2  # tariff per hour
    no_vat_row = total_row + 3  # cost without VAT
    vat_row = total_row + 4     # cost with VAT

    # Clear the template's fixed colored cells so stale content doesn't
    # appear on copied sheets.  For the data rows (last-day and total) we
    # must NOT overwrite employee columns — when there are 7+ employees,
    # column 8 (TMPL_VALUE_COL) is an employee column and clearing it would
    # destroy that employee's hours/total that _populate_template_core_sheet
    # just wrote.  The tariff-section rows (tariff_row and beyond) are safely
    # past all employee data so we always clear those.
    for r in (label_row, total_row, tariff_row, no_vat_row, vat_row):
        for c in (TMPL_VALUE_COL, TMPL_LABEL_COL):
            if r in (label_row, total_row) and c <= last_data_col:
                continue  # employee column in a data row — leave it alone
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell._style = copy(style_body)

    if hourly_tariff is not None:
        tariff = float(hourly_tariff)

        total_letters = [get_column_letter(c) for c in range(2, last_data_col + 1)]
        grand_total_formula = '=SUM(' + ','.join(f'{cl}{total_row}' for cl in total_letters) + ')'
        value_col_letter = get_column_letter(value_col)

        # Site total label (last day row, side column)
        ws.cell(row=label_row, column=value_col, value='סה"כ שעות לאתר:')._style = copy(style_site_total)

        # Grand total value (total row, side column)
        ws.cell(row=total_row, column=value_col, value=grand_total_formula)._style = copy(style_site_total)

        NIS_FORMAT = '[$₪-40D]#,##0.00'

        # Tariff per hour (numeric so no-VAT row can reference it in a formula)
        tariff_cell = ws.cell(row=tariff_row, column=value_col)
        tariff_cell.value = tariff
        tariff_cell._style = copy(style_tariff)
        tariff_cell.number_format = NIS_FORMAT
        ws.cell(row=tariff_row, column=label_col, value='מחיר לשעה')._style = copy(style_tariff_label)

        # Cost without VAT
        price_no_vat_cell = ws.cell(row=no_vat_row, column=value_col)
        price_no_vat_cell.value = f'={value_col_letter}{total_row}*{value_col_letter}{tariff_row}'
        price_no_vat_cell._style = copy(style_tariff)
        price_no_vat_cell.number_format = NIS_FORMAT
        ws.cell(row=no_vat_row, column=label_col, value='מחיר ללא מע"מ')._style = copy(style_tariff_label)

        # Cost with VAT (×1.18)
        price_vat_cell = ws.cell(row=vat_row, column=value_col)
        price_vat_cell.value = f'={value_col_letter}{no_vat_row}*1.18'
        price_vat_cell._style = copy(style_tariff)
        price_vat_cell.number_format = NIS_FORMAT
        ws.cell(row=vat_row, column=label_col, value='מחיר כולל מע"מ')._style = copy(style_tariff_label)


def _load_hours_matrix(site_id, processing_month, approved_only, include_inactive):
    started_at = time.perf_counter()
    month = datetime.strptime(processing_month, '%Y-%m-%d').date()
    site_results = load_hours_matrix_for_sites(
        site_ids=[site_id],
        processing_month=processing_month,
        approved_only=approved_only,
        include_inactive=include_inactive,
        business_id=g.business_id,
    )
    site_data = site_results.get(site_id, {'employees': [], 'matrix': {}, 'status_map': {}, 'status_matrix': {}, 'monthly_totals': {}})
    return site_data['employees'], site_data['matrix'], site_data['status_map'], month, site_data['status_matrix'], site_data['monthly_totals']


def load_hours_matrix_for_sites(site_ids, processing_month, approved_only, include_inactive, business_id):
    """Bulk load employees + best-card hours matrix for multiple sites in a fixed query budget."""
    month = datetime.strptime(processing_month, '%Y-%m-%d').date()
    unique_site_ids = list(dict.fromkeys(site_ids or []))
    if not unique_site_ids:
        return {}

    site_results = {
        site_id: {'employees': [], 'matrix': {}, 'status_map': {}, 'status_matrix': {}, 'monthly_totals': {}}
        for site_id in unique_site_ids
    }
    target_site_ids = set(unique_site_ids)

    # Baseline columns: employees whose home site is one of the target sites.
    # They appear as a column even with zero hours (preserves prior behavior).
    employee_query = db.session.query(Employee).filter(
        Employee.business_id == business_id,
        Employee.site_id.in_(unique_site_ids),
    )
    if not include_inactive:
        employee_query = employee_query.filter(Employee.is_active.is_(True))
    home_employees = employee_query.all()

    # Visiting employees: managed elsewhere (their card belongs to another site)
    # but with day entries attributed to a target site this month. Their hours
    # must surface here even though no card of theirs belongs to this site.
    visiting_employee_ids = {
        row[0]
        for row in db.session.query(WorkCard.employee_id)
        .join(WorkCardDayEntry, WorkCardDayEntry.work_card_id == WorkCard.id)
        .filter(
            WorkCard.business_id == business_id,
            WorkCard.processing_month == month,
            WorkCard.employee_id.isnot(None),
            WorkCardDayEntry.attributed_site_id.in_(unique_site_ids),
        )
        .distinct()
    }

    employees_by_id = {emp.id: emp for emp in home_employees}
    missing_ids = [eid for eid in visiting_employee_ids if eid not in employees_by_id]
    if missing_ids:
        visiting_query = db.session.query(Employee).filter(
            Employee.business_id == business_id,
            Employee.id.in_(missing_ids),
        )
        if not include_inactive:
            visiting_query = visiting_query.filter(Employee.is_active.is_(True))
        for emp in visiting_query.all():
            employees_by_id[emp.id] = emp

    # Seed home-employee columns; we add visiting employees as their entries land.
    added_columns = {site_id: set() for site_id in unique_site_ids}
    for emp in home_employees:
        site_results[emp.site_id]['employees'].append(emp)
        added_columns[emp.site_id].add(emp.id)

    def _finalize():
        for site_data in site_results.values():
            site_data['employees'] = _sort_employees_for_export(site_data['employees'])
        return site_results

    relevant_employee_ids = list({emp.id for emp in home_employees} | visiting_employee_ids)
    if not relevant_employee_ids:
        return _finalize()

    # Best (managing) card per relevant employee for the month, ranked across ALL
    # their cards regardless of site: a transferred employee's card lives at their
    # final/home site yet contributes days to the sites they moved through.
    ranked_cards = db.session.query(
        WorkCard.id.label('work_card_id'),
        WorkCard.site_id,
        WorkCard.employee_id,
        WorkCard.review_status,
        WorkCard.monthly_total_hours,
        func.row_number().over(
            partition_by=WorkCard.employee_id,
            order_by=[
                case(
                    (WorkCard.review_status == 'APPROVED', 1),
                    else_=2
                ),
                WorkCard.created_at.desc()
            ]
        ).label('rank')
    ).filter(
        WorkCard.business_id == business_id,
        WorkCard.processing_month == month,
        WorkCard.employee_id.in_(relevant_employee_ids),
    )

    if approved_only:
        ranked_cards = ranked_cards.filter(WorkCard.review_status == 'APPROVED')

    ranked_cards = ranked_cards.subquery()

    best_cards_rows = db.session.query(
        ranked_cards.c.work_card_id,
        ranked_cards.c.site_id,
        ranked_cards.c.employee_id,
        ranked_cards.c.review_status,
        ranked_cards.c.monthly_total_hours,
    ).filter(
        ranked_cards.c.rank == 1
    ).all()

    if not best_cards_rows:
        return _finalize()

    # card_id -> (card_site_id, employee_id, review_status)
    work_card_meta = {}
    work_card_ids = []
    for row in best_cards_rows:
        work_card_ids.append(row.work_card_id)
        work_card_meta[str(row.work_card_id)] = (row.site_id, row.employee_id, row.review_status)
        # The managing card's site records the employee's overall status (for the
        # summary view), if that site is in scope.
        if row.site_id in target_site_ids:
            site_results[row.site_id]['status_map'][str(row.employee_id)] = row.review_status

    day_entries = db.session.query(
        WorkCardDayEntry.work_card_id,
        WorkCardDayEntry.day_of_month,
        WorkCardDayEntry.total_hours,
        WorkCardDayEntry.day_status,
        WorkCardDayEntry.attributed_site_id,
    ).filter(
        WorkCardDayEntry.work_card_id.in_(work_card_ids)
    ).all()

    # Employees with any cross-site attribution: their hours are spread per-day,
    # so a single card-level monthly_total can't be applied (see below).
    split_employee_ids = set()

    for entry in day_entries:
        meta = work_card_meta.get(str(entry.work_card_id))
        if not meta:
            continue
        card_site_id, employee_id, review_status = meta
        employee_id_str = str(employee_id)

        effective_site = entry.attributed_site_id or card_site_id
        if entry.attributed_site_id is not None and entry.attributed_site_id != card_site_id:
            split_employee_ids.add(employee_id_str)

        if effective_site not in target_site_ids:
            continue

        site_data = site_results[effective_site]
        # Surface a visiting employee as a column the first time a day lands here.
        if employee_id not in added_columns[effective_site] and employee_id in employees_by_id:
            site_data['employees'].append(employees_by_id[employee_id])
            added_columns[effective_site].add(employee_id)
        site_data['status_map'][employee_id_str] = review_status

        if entry.day_status:
            site_data['status_matrix'].setdefault(employee_id_str, {})[entry.day_of_month] = entry.day_status
        elif entry.total_hours is not None:
            site_data['matrix'].setdefault(employee_id_str, {})[entry.day_of_month] = float(entry.total_hours)

    # monthly_total_hours is a single card-level figure used when per-day hours
    # aren't recorded. It cannot be divided across sites, so it only applies to a
    # non-split employee, attributed to their managing card's site.
    for row in best_cards_rows:
        if row.monthly_total_hours is None:
            continue
        employee_id_str = str(row.employee_id)
        if employee_id_str in split_employee_ids:
            continue
        if row.site_id in target_site_ids:
            site_results[row.site_id]['monthly_totals'][employee_id_str] = float(row.monthly_total_hours)
            if row.employee_id not in added_columns[row.site_id] and row.employee_id in employees_by_id:
                site_results[row.site_id]['employees'].append(employees_by_id[row.employee_id])
                added_columns[row.site_id].add(row.employee_id)

    return _finalize()

@sites_bp.route('', methods=['GET'])
@token_required
def get_sites():
    """Get all sites, optionally with employee counts, scoped to tenant."""
    try:
        include_counts = request.args.get('include_counts', 'false').lower() == 'true'
        only_active = request.args.get('active', 'false').lower() == 'true'
        
        # Always scope to current business
        business_id = g.business_id
        
        if include_counts:
            if only_active:
                results = repo.get_active_with_employee_count(business_id)
            else:
                results = repo.get_with_employee_count(business_id)
                
            data = []
            for item in results:
                site_dict = model_to_dict(item['site'])
                site_dict['employee_count'] = item['employee_count']
                data.append(site_dict)
        else:
            if only_active:
                sites = repo.get_active_sites(business_id)
            else:
                sites = repo.get_all(filters={'business_id': business_id})
            data = models_to_list(sites)
            
        return api_response(data=data)
    except Exception as e:
        logger.exception("Failed to get sites")
        traceback.print_exc()
        return api_response(status_code=500, message="Failed to get sites", error=str(e))

def _coerce_expected_cards(data):
    """Validate/normalize expected_work_cards_per_month in-place. Returns error msg or None."""
    if 'expected_work_cards_per_month' not in data:
        return None
    val = data.get('expected_work_cards_per_month')
    if val in (None, ''):
        data['expected_work_cards_per_month'] = None
        return None
    try:
        ival = int(val)
    except (TypeError, ValueError):
        return "מספר כרטיסים צפוי חייב להיות מספר שלם"
    if not (1 <= ival <= 10):
        return "מספר כרטיסים צפוי חייב להיות בין 1 ל-10"
    data['expected_work_cards_per_month'] = ival
    return None


@sites_bp.route('', methods=['POST'])
@token_required
@role_required('ADMIN')
def create_site():
    """Create a new site, scoped to tenant."""
    data = request.get_json()
    if not data:
        return api_response(status_code=400, message="No data provided", error="Bad Request")
        
    try:
        # Enforce tenant scoping
        data['business_id'] = g.business_id
        
        # Validation
        if not data.get('site_name'):
            return api_response(status_code=400, message="Site name is required", error="Bad Request")
        
        # Check if site with name exists within tenant
        existing = repo.get_by_name_and_business(data['site_name'], g.business_id)
        if existing:
             return api_response(status_code=409, message="Site with this name already exists", error="Conflict")

        if 'contractor_emails' in data:
            emails, err = _validate_contractor_emails(data)
            if err:
                return api_response(status_code=400, message=err, error="Bad Request")
            data['contractor_emails'] = emails

        if 'contractor_phone_number' in data:
            phone, err = _normalize_contractor_phone(data.get('contractor_phone_number'))
            if err:
                return api_response(status_code=400, message=err, error="Bad Request")
            data['contractor_phone_number'] = phone

        err = _coerce_expected_cards(data)
        if err:
            return api_response(status_code=400, message=err, error="Bad Request")

        site = repo.create(**data)
        invalidate_business_cache(g.business_id)
        return api_response(data=model_to_dict(site), message="Site created successfully", status_code=201)
    except Exception as e:
        logger.exception("Failed to create site")
        traceback.print_exc()
        return api_response(status_code=500, message="Failed to create site", error=str(e))

@sites_bp.route('/<uuid:site_id>', methods=['GET'])
@token_required
def get_site(site_id):
    """Get a specific site by ID, scoped to tenant."""
    try:
        site = repo.get_by_id(site_id)
        if not site or site.business_id != g.business_id:
            return api_response(status_code=404, message="Site not found", error="Not Found")
            
        return api_response(data=model_to_dict(site))
    except Exception as e:
        logger.exception(f"Failed to get site {site_id}")
        traceback.print_exc()
        return api_response(status_code=500, message="Failed to get site", error=str(e))

@sites_bp.route('/<uuid:site_id>', methods=['PUT'])
@token_required
@role_required('ADMIN')
def update_site(site_id):
    """Update a site, scoped to tenant."""
    data = request.get_json()
    if not data:
        return api_response(status_code=400, message="No data provided", error="Bad Request")
        
    try:
        # Verify site belongs to user's business
        site = repo.get_by_id(site_id)
        if not site or site.business_id != g.business_id:
            return api_response(status_code=404, message="Site not found", error="Not Found")
        
        # Don't allow changing business_id
        data.pop('business_id', None)

        if 'contractor_emails' in data:
            emails, err = _validate_contractor_emails(data)
            if err:
                return api_response(status_code=400, message=err, error="Bad Request")
            data['contractor_emails'] = emails

        if 'contractor_phone_number' in data:
            phone, err = _normalize_contractor_phone(data.get('contractor_phone_number'))
            if err:
                return api_response(status_code=400, message=err, error="Bad Request")
            data['contractor_phone_number'] = phone

        if 'hourly_tariff' in data:
            val = data.get('hourly_tariff')
            if val is not None and float(val) < 0:
                return api_response(status_code=400, message="תעריף לשעה חייב להיות חיובי", error="Bad Request")

        if 'responsible_employee_id' in data:
            responsible_employee_id = data.get('responsible_employee_id')
            if not responsible_employee_id:
                data['responsible_employee_id'] = None
            else:
                try:
                    responsible_employee_uuid = uuid.UUID(str(responsible_employee_id))
                except ValueError:
                    return api_response(status_code=400, message="Invalid responsible_employee_id format", error="Bad Request")

                employee = employee_repo.get_by_id(responsible_employee_uuid)
                if not employee or employee.business_id != g.business_id or str(employee.site_id) != str(site_id):
                    return api_response(status_code=404, message="Responsible employee not found for this site", error="Not Found")
                if not employee.is_active:
                    return api_response(status_code=400, message="Responsible employee is not active", error="Bad Request")

                data['responsible_employee_id'] = responsible_employee_uuid

        if 'field_manager_id' in data:
            field_manager_id = data.get('field_manager_id')
            if not field_manager_id:
                data['field_manager_id'] = None
            else:
                try:
                    field_manager_uuid = uuid.UUID(str(field_manager_id))
                except ValueError:
                    return api_response(status_code=400, message="Invalid field_manager_id format", error="Bad Request")

                manager = user_repo.get_by_id(field_manager_uuid)
                if not manager or manager.business_id != g.business_id or manager.role != 'FIELD_MANAGER':
                    return api_response(status_code=404, message="Field manager not found for this business", error="Not Found")

                data['field_manager_id'] = field_manager_uuid

        err = _coerce_expected_cards(data)
        if err:
            return api_response(status_code=400, message=err, error="Bad Request")

        updated_site = repo.update(site_id, **data)
        if not updated_site:
            return api_response(status_code=404, message="Site not found", error="Not Found")

        invalidate_business_cache(g.business_id)
        return api_response(data=model_to_dict(updated_site), message="Site updated successfully")
    except Exception as e:
        logger.exception(f"Failed to update site {site_id}")
        traceback.print_exc()
        return api_response(status_code=500, message="Failed to update site", error=str(e))

@sites_bp.route('/<uuid:site_id>', methods=['DELETE'])
@token_required
@role_required('ADMIN')
def delete_site(site_id):
    """Delete a site, scoped to tenant."""
    try:
        # Verify site belongs to user's business
        site = repo.get_by_id(site_id)
        if not site or site.business_id != g.business_id:
            return api_response(status_code=404, message="Site not found", error="Not Found")
        
        # We might want to just mark as inactive instead of deleting if there are related records
        # But BaseRepository.delete does a hard delete. 
        # For now, let's assume hard delete is requested, but catch FK errors
        success = repo.delete(site_id)
        if not success:
            return api_response(status_code=404, message="Site not found", error="Not Found")

        invalidate_business_cache(g.business_id)
        return api_response(message="Site deleted successfully")
    except Exception as e:
        logger.exception(f"Failed to delete site {site_id}")
        traceback.print_exc()
        return api_response(status_code=500, message="Failed to delete site", error=str(e))

@sites_bp.route('/<uuid:site_id>/employee-upload-status', methods=['GET'])
@token_required
def get_employee_upload_status(site_id):
    """Get employee upload status for a site and month."""
    # Verify site belongs to user's business
    site = repo.get_by_id(site_id)
    if not site or site.business_id != g.business_id:
        return api_response(status_code=404, message="Site not found", error="Not Found")
    
    processing_month = request.args.get('processing_month')
    if not processing_month:
        return api_response(status_code=400, message="processing_month is required", error="Bad Request")
    
    try:
        # Parse processing_month
        month = datetime.strptime(processing_month, '%Y-%m-%d').date()
        
        employee_rows = get_latest_work_card_with_extraction_by_employee(
            business_id=g.business_id,
            site_id=site_id,
            processing_month=month,
        )
        status_map = build_employee_upload_status_map(employee_rows)

        result = []
        for employee, _, _, _ in employee_rows:
            employee_status = status_map.get(str(employee.id), {'status': 'NO_UPLOAD', 'work_card_id': None})
            employee_dict = model_to_dict(employee)
            result.append({
                'employee': employee_dict,
                'status': employee_status['status'],
                'work_card_id': employee_status['work_card_id'],
            })
        
        return api_response(data=result)
    except ValueError as e:
        return api_response(status_code=400, message="Invalid date format. Use YYYY-MM-DD", error=str(e))
    except Exception as e:
        logger.exception(f"Failed to get employee upload status for site {site_id}")
        traceback.print_exc()
        return api_response(status_code=500, message="Failed to get employee upload status", error=str(e))

@sites_bp.route('/<uuid:site_id>/matrix', methods=['GET'])
@token_required
def get_hours_matrix(site_id):
    """Get hours matrix for a site and month with performance optimization."""
    started_at = time.perf_counter()
    with QueryCounter(db.engine) as query_counter:
        # Verify site belongs to user's business
        site = repo.get_by_id(site_id)
        if not site or site.business_id != g.business_id:
            return api_response(status_code=404, message="Site not found", error="Not Found")

        processing_month = request.args.get('processing_month')
        if not processing_month:
            return api_response(status_code=400, message="processing_month is required", error="Bad Request")

        approved_only = request.args.get('approved_only', 'true').lower() == 'true'
        include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'

        try:
            employees, matrix, status_map, _, status_matrix, monthly_totals = _load_hours_matrix(
                site_id,
                processing_month,
                approved_only,
                include_inactive
            )

            response = api_response(data={
                'employees': models_to_list(employees),
                'matrix': matrix,
                'status_map': status_map,
                'status_matrix': status_matrix,
                'monthly_totals': monthly_totals,
            })
            return response
        except ValueError as e:
            return api_response(status_code=400, message="Invalid date format. Use YYYY-MM-DD", error=str(e))
        except Exception as e:
            logger.exception(f"Failed to get hours matrix for site {site_id}")
            traceback.print_exc()
            return api_response(status_code=500, message="Failed to get hours matrix", error=str(e))
        finally:
            duration_ms = round((time.perf_counter() - started_at) * 1000, 2)
            logger.info(
                "sites.matrix_route",
                extra={
                    'duration_ms': duration_ms,
                    'site_count': 1,
                    'employee_count': len(employees) if 'employees' in locals() else 0,
                    'query_count': query_counter.count,
                    'output_type': 'json',
                }
            )

def _generate_summary_xlsx(site, employees, matrix, month, status_matrix, monthly_totals=None):
    """Generate a monthly summary XLSX workbook and return (BytesIO, filename)."""
    template_path = _resolve_summary_template_path()
    workbook = load_workbook(template_path)

    ws = workbook.worksheets[0]
    style_header = copy(ws['B1']._style)
    style_body = copy(ws['B3']._style)
    style_total = copy(ws['A34']._style)
    style_site_total = copy(ws.cell(row=33, column=8)._style)
    style_tariff = copy(ws.cell(row=36, column=8)._style)
    style_tariff_label = copy(ws.cell(row=36, column=9)._style)

    ws.title = _safe_sheet_name(site.site_name, set())

    total_row = _populate_template_core_sheet(
        ws,
        employees,
        matrix,
        month,
        style_header=style_header,
        style_body=style_body,
        style_total=style_total,
        status_matrix=status_matrix,
        monthly_totals=monthly_totals,
    )

    for extra_ws in workbook.worksheets[1:]:
        workbook.remove(extra_ws)

    _add_tariff_summary(
        ws,
        employee_count=len(employees),
        hourly_tariff=site.hourly_tariff,
        style_body=style_body,
        style_site_total=style_site_total,
        style_tariff=style_tariff,
        style_tariff_label=style_tariff_label,
        total_row=total_row,
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    site_label = _safe_label(site.site_name) or str(site.id)
    filename = f"monthly_summary_{site.id}_{site_label}_{month.strftime('%Y-%m')}.xlsx"
    return output, filename


@sites_bp.route('/<uuid:site_id>/summary/export', methods=['GET'])
@token_required
@role_required('ADMIN')
def export_monthly_summary(site_id):
    """Export monthly summary matrix as a CSV file."""
    site = repo.get_by_id(site_id)
    if not site or site.business_id != g.business_id:
        return api_response(status_code=404, message="Site not found", error="Not Found")

    processing_month = request.args.get('processing_month')
    if not processing_month:
        return api_response(status_code=400, message="processing_month is required", error="Bad Request")

    approved_only = request.args.get('approved_only', 'false').lower() == 'true'
    include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'

    try:
        employees, matrix, status_map, month, status_matrix, monthly_totals = _load_hours_matrix(
            site_id,
            processing_month,
            approved_only,
            include_inactive
        )
    except ValueError as e:
        return api_response(status_code=400, message="Invalid date format. Use YYYY-MM-DD", error=str(e))
    except Exception as e:
        logger.exception(f"Failed to export hours matrix for site {site_id}")
        traceback.print_exc()
        return api_response(status_code=500, message="Failed to export summary", error=str(e))

    try:
        output, download_name = _generate_summary_xlsx(site, employees, matrix, month, status_matrix, monthly_totals)
    except Exception as e:
        logger.exception(f"Failed to generate summary XLSX for site {site_id}")
        return api_response(status_code=500, message="Failed to generate summary", error=str(e))

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name
    )

@sites_bp.route('/<uuid:site_id>/summary/email', methods=['POST'])
@token_required
@role_required('ADMIN')
def send_summary_email(site_id):
    """Send the monthly summary XLSX as an email attachment to the site's contractor emails."""
    site = repo.get_by_id(site_id)
    if not site or site.business_id != g.business_id:
        return api_response(status_code=404, message="Site not found", error="Not Found")

    recipients = site.contractor_emails or []
    if not recipients:
        return api_response(
            status_code=400,
            message="לא הוגדרו כתובות מייל לאתר זה",
            error="No email recipients configured",
        )

    data = request.get_json() or {}
    processing_month = data.get('processing_month')
    if not processing_month:
        return api_response(status_code=400, message="processing_month is required", error="Bad Request")

    try:
        employees, matrix, _, month, status_matrix, monthly_totals = _load_hours_matrix(
            site_id, processing_month, approved_only=False, include_inactive=False
        )
    except ValueError as e:
        return api_response(status_code=400, message="Invalid date format. Use YYYY-MM-DD", error=str(e))
    except Exception as e:
        logger.exception(f"Failed to load hours matrix for email send, site {site_id}")
        return api_response(status_code=500, message="Failed to generate summary", error=str(e))

    try:
        output, filename = _generate_summary_xlsx(site, employees, matrix, month, status_matrix, monthly_totals)
    except Exception as e:
        logger.exception(f"Failed to generate summary XLSX for email, site {site_id}")
        return api_response(status_code=500, message="Failed to generate summary", error=str(e))

    month_label = month.strftime('%Y-%m')
    subject = f"סיכום חודשי - {site.site_name} - {month_label}"
    html_body = f"""
    <div dir="rtl" style="font-family: Arial, sans-serif;">
        <p>שלום,</p>
        <p>מצורף סיכום שעות חודשי עבור אתר <strong>{site.site_name}</strong> לחודש <strong>{month_label}</strong>.</p>
        <br>
        <p>בברכה,<br>AutoHQ</p>
    </div>
    """

    try:
        result = send_email_with_attachment(
            to=recipients,
            subject=subject,
            html_body=html_body,
            attachment_bytes=output.read(),
            attachment_filename=filename,
        )
    except ValueError as e:
        logger.error(f"Email config error: {e}")
        return api_response(status_code=500, message="Email service not configured", error=str(e))
    except Exception as e:
        logger.exception(f"Failed to send summary email for site {site_id}")
        return api_response(status_code=500, message="שגיאה בשליחת המייל", error=str(e))

    return api_response(
        data={
            "status": "sent",
            "recipients": recipients,
            "site_name": site.site_name,
        },
        message=f"הסיכום נשלח בהצלחה ל-{len(recipients)} נמענים",
    )


@sites_bp.route('/<uuid:site_id>/summary/whatsapp', methods=['POST'])
@token_required
@role_required('ADMIN')
def send_summary_whatsapp(site_id):
    """Send the monthly summary XLSX as a WhatsApp document to the site's contractor phone number."""
    site = repo.get_by_id(site_id)
    if not site or site.business_id != g.business_id:
        return api_response(status_code=404, message="Site not found", error="Not Found")

    phone = site.contractor_phone_number
    if not phone:
        return api_response(
            status_code=400,
            message="לא הוגדר מספר טלפון לאתר זה",
            error="No contractor phone number configured",
        )

    data = request.get_json() or {}
    processing_month = data.get('processing_month')
    if not processing_month:
        return api_response(status_code=400, message="processing_month is required", error="Bad Request")

    client = WhatsAppListenerClient.from_env()
    if client is None:
        return api_response(
            status_code=503,
            message="שירות הוואטסאפ לא מוגדר",
            error="WA_LISTENER_URL is not set",
        )

    try:
        employees, matrix, _, month, status_matrix, monthly_totals = _load_hours_matrix(
            site_id, processing_month, approved_only=False, include_inactive=False
        )
    except ValueError as e:
        return api_response(status_code=400, message="Invalid date format. Use YYYY-MM-DD", error=str(e))
    except Exception as e:
        logger.exception(f"Failed to load hours matrix for WhatsApp send, site {site_id}")
        return api_response(status_code=500, message="Failed to generate summary", error=str(e))

    try:
        output, _ = _generate_summary_xlsx(site, employees, matrix, month, status_matrix, monthly_totals)
    except Exception as e:
        logger.exception(f"Failed to generate summary XLSX for WhatsApp, site {site_id}")
        return api_response(status_code=500, message="Failed to generate summary", error=str(e))

    month_label = month.strftime('%Y-%m')
    business = business_repo.get_by_id(site.business_id)
    business_name = business.name if business else ''
    filename = f"סיכום חודשי - {site.site_name} - {month_label}.xlsx"
    caption = (
        f"שלום,\n"
        f"מצורף דוח סיכום שעות חודשי לאתר {site.site_name} עבור חודש {month_label}.\n"
        f"בברכה,\n"
        f"{business_name}"
    )
    chat_id = f"{phone}@s.whatsapp.net"

    try:
        client.send_document(
            chat_id=chat_id,
            file_bytes=output.read(),
            filename=filename,
            caption=caption,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except WhatsAppNumberNotRegisteredError:
        return api_response(
            status_code=404,
            message="המספר אינו רשום בוואטסאפ",
            error="Phone number is not registered on WhatsApp",
        )
    except WhatsAppNotConnectedError:
        return api_response(
            status_code=503,
            message="שירות הוואטסאפ אינו מחובר כרגע",
            error="WhatsApp listener not connected",
        )
    except WhatsAppPayloadTooLargeError:
        return api_response(
            status_code=413,
            message="הקובץ גדול מדי לשליחה בוואטסאפ",
            error="XLSX file exceeds WhatsApp size limit",
        )
    except WhatsAppAuthError as e:
        logger.error(f"WhatsApp listener auth error: {e}")
        return api_response(status_code=500, message="שגיאת הזדהות מול שירות הוואטסאפ", error=str(e))
    except WhatsAppListenerError as e:
        logger.exception(f"WhatsApp listener error sending summary for site {site_id}")
        return api_response(status_code=502, message="שגיאה בשליחת הוואטסאפ", error=str(e))

    return api_response(
        data={
            "status": "sent",
            "phone": phone,
            "site_name": site.site_name,
        },
        message="הסיכום נשלח בוואטסאפ בהצלחה",
    )


@sites_bp.route('/summary/export-batch', methods=['GET'])
@token_required
@role_required('ADMIN')
def export_monthly_summary_batch():
    """Export monthly summary matrix for all sites in template format (one sheet per site)."""
    processing_month = request.args.get('processing_month')
    if not processing_month:
        return api_response(status_code=400, message="processing_month is required", error="Bad Request")

    approved_only = request.args.get('approved_only', 'false').lower() == 'true'
    include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'
    include_inactive_sites = request.args.get('include_inactive_sites', 'false').lower() == 'true'

    try:
        month = datetime.strptime(processing_month, '%Y-%m-%d').date()
    except ValueError as e:
        return api_response(status_code=400, message="Invalid date format. Use YYYY-MM-DD", error=str(e))

    started_at = time.perf_counter()
    total_employee_count = 0
    with QueryCounter(db.engine) as query_counter:
        sites = repo.get_all_for_business(g.business_id)
        if not include_inactive_sites:
            sites = [site for site in sites if site.is_active]
        sites = sorted(
            sites,
            key=lambda s: (
                (s.site_name or '').strip().lower(),
                (s.site_code or '').strip().lower(),
                str(s.id)
            )
        )

        try:
            template_path = _resolve_summary_template_path()
            workbook = load_workbook(template_path)
        except Exception as e:
            logger.exception("Failed to load summary export template")
            return api_response(status_code=500, message="Failed to load export template", error=str(e))

        template_ws = workbook.worksheets[0]
        style_header = copy(template_ws['B1']._style)
        style_body = copy(template_ws['B3']._style)
        style_total = copy(template_ws['A34']._style)
        style_site_total = copy(template_ws.cell(row=33, column=8)._style)
        style_tariff = copy(template_ws.cell(row=36, column=8)._style)
        style_tariff_label = copy(template_ws.cell(row=36, column=9)._style)

        for ws in workbook.worksheets[1:]:
            workbook.remove(ws)

        site_matrices = load_hours_matrix_for_sites(
            site_ids=[site.id for site in sites],
            processing_month=processing_month,
            approved_only=approved_only,
            include_inactive=include_inactive,
            business_id=g.business_id,
        )

        used_sheet_names = set()
        for site in sites:
            site_data = site_matrices.get(site.id, {'employees': [], 'matrix': {}, 'status_map': {}, 'status_matrix': {}, 'monthly_totals': {}})
            total_employee_count += len(site_data['employees'])

            ws = workbook.copy_worksheet(template_ws)
            ws.title = _safe_sheet_name(site.site_name, used_sheet_names)
            total_row = _populate_template_core_sheet(
                ws,
                site_data['employees'],
                site_data['matrix'],
                month,
                style_header=style_header,
                style_body=style_body,
                style_total=style_total,
                status_matrix=site_data['status_matrix'],
                monthly_totals=site_data.get('monthly_totals'),
            )
            _add_tariff_summary(
                ws,
                employee_count=len(site_data['employees']),
                hourly_tariff=site.hourly_tariff,
                style_body=style_body,
                style_site_total=style_site_total,
                style_tariff=style_tariff,
                style_tariff_label=style_tariff_label,
                total_row=total_row,
            )

        if workbook.worksheets and len(workbook.worksheets) > 1:
            workbook.remove(template_ws)
        else:
            _populate_template_core_sheet(
                template_ws,
                [],
                {},
                month,
                style_header=style_header,
                style_body=style_body,
                style_total=style_total,
            )
            template_ws.title = 'Sites'

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

    duration_s = time.perf_counter() - started_at
    sites_metrics.observe_export_generation_latency(duration_s, output_type='xlsx', endpoint='summary_export_batch')
    logger.info(
        "sites.export_batch_route",
        extra={
            'duration_ms': round(duration_s * 1000, 2),
            'site_count': len(sites),
            'employee_count': total_employee_count,
            'query_count': query_counter.count,
            'output_type': 'xlsx',
        }
    )

    download_name = f"monthly_summary_all_sites_{month.strftime('%Y-%m')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name
    )

@sites_bp.route('/<uuid:site_id>/salary-template/export', methods=['GET'])
@token_required
@role_required('ADMIN')
def export_salary_template_site(site_id):
    """Export salary template workbook for a single site."""
    site = repo.get_by_id(site_id)
    if not site or site.business_id != g.business_id:
        return api_response(status_code=404, message="Site not found", error="Not Found")

    processing_month = request.args.get('processing_month')
    if not processing_month:
        return api_response(status_code=400, message="processing_month is required", error="Bad Request")

    include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'

    try:
        month = datetime.strptime(processing_month, '%Y-%m-%d').date()
    except ValueError as e:
        return api_response(status_code=400, message="Invalid date format. Use YYYY-MM-DD", error=str(e))

    try:
        employees, matrix, _, _, status_matrix, __ = _load_hours_matrix(
            site_id,
            processing_month,
            approved_only=False,
            include_inactive=include_inactive
        )

        template_path = _resolve_salary_template_path(month)
        workbook = load_workbook(template_path)
        ws = workbook.worksheets[0]
        _populate_salary_template_sheet(ws, employees, matrix, month, status_matrix)
        ws.title = _safe_sheet_name(site.site_name, set())
    except ValueError as e:
        return api_response(status_code=500, message="Invalid salary template format", error=str(e))
    except Exception as e:
        logger.exception(f"Failed to export salary template for site {site_id}")
        return api_response(status_code=500, message="Failed to export salary template", error=str(e))

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    site_label = _safe_label(site.site_name) or str(site.id)
    download_name = f"salary_template_{site_label}_{month.strftime('%Y-%m')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name
    )

@sites_bp.route('/salary-template/export-batch', methods=['GET'])
@token_required
@role_required('ADMIN')
def export_salary_template_batch():
    """Export salary template workbook for all sites (one sheet per site)."""
    processing_month = request.args.get('processing_month')
    if not processing_month:
        return api_response(status_code=400, message="processing_month is required", error="Bad Request")

    include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'
    include_inactive_sites = request.args.get('include_inactive_sites', 'false').lower() == 'true'

    try:
        month = datetime.strptime(processing_month, '%Y-%m-%d').date()
    except ValueError as e:
        return api_response(status_code=400, message="Invalid date format. Use YYYY-MM-DD", error=str(e))

    started_at = time.perf_counter()
    total_employee_count = 0
    with QueryCounter(db.engine) as query_counter:
        sites = repo.get_all_for_business(g.business_id)
        if not include_inactive_sites:
            sites = [site for site in sites if site.is_active]
        sites = sorted(
            sites,
            key=lambda s: (
                (s.site_name or '').strip().lower(),
                (s.site_code or '').strip().lower(),
                str(s.id)
            )
        )

        try:
            template_path = _resolve_salary_template_path(month)
            workbook = load_workbook(template_path)
        except Exception as e:
            logger.exception("Failed to load salary export template")
            return api_response(status_code=500, message="Failed to load salary template", error=str(e))

        template_ws = workbook.worksheets[0]
        for ws in workbook.worksheets[1:]:
            workbook.remove(ws)

        site_matrices = load_hours_matrix_for_sites(
            site_ids=[site.id for site in sites],
            processing_month=processing_month,
            approved_only=False,
            include_inactive=include_inactive,
            business_id=g.business_id,
        )

        used_sheet_names = set()
        populated_count = 0
        for site in sites:
            site_data = site_matrices.get(site.id, {'employees': [], 'matrix': {}, 'status_map': {}, 'status_matrix': {}})
            total_employee_count += len(site_data['employees'])

            ws = workbook.copy_worksheet(template_ws)
            ws.title = _safe_sheet_name(site.site_name, used_sheet_names)
            try:
                _populate_salary_template_sheet(ws, site_data['employees'], site_data['matrix'], month, site_data['status_matrix'])
                populated_count += 1
            except ValueError as e:
                logger.exception(f"Invalid salary template format for site {site.id}")
                return api_response(status_code=500, message="Invalid salary template format", error=str(e))

        if populated_count > 0 and len(workbook.worksheets) > 1:
            workbook.remove(template_ws)
        else:
            try:
                _populate_salary_template_sheet(template_ws, [], {}, month)
                template_ws.title = 'Sites'
            except ValueError as e:
                return api_response(status_code=500, message="Invalid salary template format", error=str(e))

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

    duration_s = time.perf_counter() - started_at
    sites_metrics.observe_export_generation_latency(duration_s, output_type='xlsx', endpoint='salary_export_batch')
    logger.info(
        "sites.salary_export_batch_route",
        extra={
            'duration_ms': round(duration_s * 1000, 2),
            'site_count': len(sites),
            'employee_count': total_employee_count,
            'query_count': query_counter.count,
            'output_type': 'xlsx',
        }
    )

    download_name = f"salary_template_all_sites_{month.strftime('%Y-%m')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name
    )


@sites_bp.route('/<uuid:site_id>/access-link', methods=['POST'])
@token_required
@role_required('ADMIN', 'OPERATOR_MANAGER')
def create_access_link(site_id):
    data = request.get_json() or {}
    employee_id = data.get('employee_id')
    processing_month = data.get('processing_month')

    if not employee_id or not processing_month:
        return api_response(status_code=400, message="employee_id and processing_month are required", error="Bad Request")

    site = repo.get_by_id(site_id)
    if not site or site.business_id != g.business_id:
        return api_response(status_code=404, message="Site not found", error="Not Found")

    employee = employee_repo.get_by_id(employee_id)
    if not employee or employee.business_id != g.business_id or str(employee.site_id) != str(site_id):
        return api_response(status_code=404, message="Employee not found for this site", error="Not Found")
    if not employee.is_active:
        return api_response(status_code=400, message="Employee is not active", error="Bad Request")

    try:
        month = datetime.strptime(processing_month, '%Y-%m-%d').date()
    except ValueError as e:
        return api_response(status_code=400, message="Invalid date format. Use YYYY-MM-DD", error=str(e))

    # Generate unique token
    token = _generate_access_token()
    if not token:
        return api_response(status_code=500, message="Failed to generate access token", error="Server Error")

    expires_at = datetime.now(timezone.utc) + timedelta(days=30)

    access_request = access_repo.create(
        token=token,
        business_id=g.business_id,
        site_id=site_id,
        employee_id=employee_id,
        processing_month=month,
        created_by_user_id=g.current_user.id,
        expires_at=expires_at,
        is_active=True
    )

    url = _build_access_link_url(token)
    data = model_to_dict(access_request)
    data['url'] = url
    data['employee_name'] = employee.full_name

    return api_response(data=data, message="Access link created", status_code=201)


@sites_bp.route('/<uuid:site_id>/access-links', methods=['GET'])
@token_required
@role_required('ADMIN', 'OPERATOR_MANAGER')
def list_access_links(site_id):
    site = repo.get_by_id(site_id)
    if not site or site.business_id != g.business_id:
        return api_response(status_code=404, message="Site not found", error="Not Found")

    links_with_employees = access_repo.list_active_for_site_with_employee(site_id, g.business_id)
    data = []
    for link, employee_name in links_with_employees:
        link_dict = model_to_dict(link)
        link_dict['employee_name'] = employee_name or ''
        link_dict['url'] = f"{request.host_url.rstrip('/')}/portal/{link.token}"
        data.append(link_dict)

    return api_response(data=data)


@sites_bp.route('/<uuid:site_id>/access-link/<uuid:request_id>/whatsapp', methods=['POST'])
@token_required
@role_required('ADMIN', 'OPERATOR_MANAGER')
def send_whatsapp_link(site_id, request_id):
    """Send an access link via WhatsApp to the employee."""
    site = repo.get_by_id(site_id)
    if not site or site.business_id != g.business_id:
        return api_response(status_code=404, message="Site not found", error="Not Found")

    access_request = access_repo.get_by_id(request_id)
    if not access_request or access_request.business_id != g.business_id:
        return api_response(status_code=404, message="Access link not found", error="Not Found")

    employee = employee_repo.get_by_id(access_request.employee_id)
    if not employee or not employee.phone_number:
        return api_response(status_code=400, message="Employee has no phone number", error="Bad Request")

    raw_phone = normalize_phone(employee.phone_number)
    if not raw_phone:
        return api_response(status_code=400, message="Invalid phone number format", error="Bad Request")

    formatted_phone = _format_whatsapp_number(raw_phone)
    if not formatted_phone:
        return api_response(status_code=400, message="Invalid phone number format", error="Bad Request")

    try:
        account_sid = os.environ.get('TWILIO_ACCOUNT_SID')
        auth_token = os.environ.get('TWILIO_AUTH_TOKEN')
        from_number = os.environ.get('TWILIO_WHATSAPP_NUMBER')

        if not all([account_sid, auth_token, from_number]):
            logger.error("Twilio credentials missing")
            return api_response(status_code=500, message="Server configuration error", error="Twilio config missing")

        client = Client(account_sid, auth_token)

        url = _build_access_link_url(access_request.token)
        message_body = (
            f"שלום {employee.full_name},\n"
            f"להלן הקישור להעלאת כרטיסי העבודה עבור חודש {access_request.processing_month.strftime('%m/%Y')}:\n"
            f"{url}"
        )

        message = client.messages.create(
            from_=from_number,
            body=message_body,
            to=f"whatsapp:{formatted_phone}"
        )

        logger.info(f"WhatsApp sent to {formatted_phone}: {message.sid}")
        return api_response(message="WhatsApp sent successfully")

    except Exception as e:
        logger.exception(f"Twilio error for request {request_id}")
        return api_response(status_code=500, message="Failed to send WhatsApp message", error=str(e))


@sites_bp.route('/access-links/whatsapp-batch', methods=['POST'])
@token_required
@role_required('ADMIN')
def send_whatsapp_links_batch():
    """Create access links and send them via WhatsApp for multiple sites."""
    payload = request.get_json() or {}
    site_ids = payload.get('site_ids') or []
    processing_month = payload.get('processing_month')

    if not site_ids or not isinstance(site_ids, list):
        return api_response(status_code=400, message="site_ids must be a non-empty list", error="Bad Request")
    if not processing_month:
        return api_response(status_code=400, message="processing_month is required", error="Bad Request")

    try:
        month = datetime.strptime(processing_month, '%Y-%m-%d').date()
    except ValueError as e:
        return api_response(status_code=400, message="Invalid date format. Use YYYY-MM-DD", error=str(e))

    account_sid = os.environ.get('TWILIO_ACCOUNT_SID')
    auth_token = os.environ.get('TWILIO_AUTH_TOKEN')
    from_number = os.environ.get('TWILIO_WHATSAPP_NUMBER')
    if not all([account_sid, auth_token, from_number]):
        logger.error("Twilio credentials missing")
        return api_response(status_code=500, message="Server configuration error", error="Twilio config missing")

    client = Client(account_sid, auth_token)

    results = []
    sent_count = 0
    failed_count = 0
    skipped_count = 0

    parsed_site_ids = []
    for site_id in site_ids:
        site_id_str = str(site_id)
        try:
            parsed_site_ids.append((site_id_str, uuid.UUID(site_id_str)))
        except ValueError:
            skipped_count += 1
            sites_metrics.increment_whatsapp_batch_outcome('skipped')
            results.append({
                'site_id': site_id_str,
                'status': 'skipped',
                'reason': 'Invalid site_id format'
            })

    site_lookup = {
        str(site.id): site
        for site in repo.get_by_ids_for_business(
            [parsed_id for _, parsed_id in parsed_site_ids],
            g.business_id,
        )
    }
    responsible_employee_ids = [
        site.responsible_employee_id
        for site in site_lookup.values()
        if site.responsible_employee_id
    ]
    employee_lookup = {
        str(employee.id): employee
        for employee in employee_repo.get_by_ids_for_business(responsible_employee_ids, g.business_id)
    }

    for site_id_str, site_uuid in parsed_site_ids:
        site = site_lookup.get(str(site_uuid))
        if not site:
            skipped_count += 1
            sites_metrics.increment_whatsapp_batch_outcome('skipped')
            results.append({
                'site_id': site_id_str,
                'status': 'skipped',
                'reason': 'Site not found'
            })
            continue

        if not site.responsible_employee_id:
            skipped_count += 1
            sites_metrics.increment_whatsapp_batch_outcome('skipped')
            results.append({
                'site_id': site_id_str,
                'site_name': site.site_name,
                'status': 'skipped',
                'reason': 'No responsible employee'
            })
            continue

        employee = employee_lookup.get(str(site.responsible_employee_id))
        if not employee or employee.business_id != g.business_id or str(employee.site_id) != str(site.id):
            skipped_count += 1
            sites_metrics.increment_whatsapp_batch_outcome('skipped')
            results.append({
                'site_id': site_id_str,
                'site_name': site.site_name,
                'employee_id': str(site.responsible_employee_id),
                'status': 'skipped',
                'reason': 'Responsible employee not found for site'
            })
            continue

        if not employee.is_active:
            skipped_count += 1
            sites_metrics.increment_whatsapp_batch_outcome('skipped')
            results.append({
                'site_id': site_id_str,
                'site_name': site.site_name,
                'employee_id': str(employee.id),
                'employee_name': employee.full_name,
                'status': 'skipped',
                'reason': 'Responsible employee is not active'
            })
            continue

        if not employee.phone_number:
            skipped_count += 1
            sites_metrics.increment_whatsapp_batch_outcome('skipped')
            results.append({
                'site_id': site_id_str,
                'site_name': site.site_name,
                'employee_id': str(employee.id),
                'employee_name': employee.full_name,
                'status': 'skipped',
                'reason': 'Employee has no phone number'
            })
            continue

        raw_phone = normalize_phone(employee.phone_number)
        if not raw_phone:
            skipped_count += 1
            sites_metrics.increment_whatsapp_batch_outcome('skipped')
            results.append({
                'site_id': site_id_str,
                'site_name': site.site_name,
                'employee_id': str(employee.id),
                'employee_name': employee.full_name,
                'status': 'skipped',
                'reason': 'Invalid phone number format'
            })
            continue

        formatted_phone = _format_whatsapp_number(raw_phone)
        if not formatted_phone:
            skipped_count += 1
            sites_metrics.increment_whatsapp_batch_outcome('skipped')
            results.append({
                'site_id': site_id_str,
                'site_name': site.site_name,
                'employee_id': str(employee.id),
                'employee_name': employee.full_name,
                'status': 'skipped',
                'reason': 'Invalid phone number format'
            })
            continue

        token = _generate_access_token()
        if not token:
            failed_count += 1
            sites_metrics.increment_whatsapp_batch_outcome('failed')
            results.append({
                'site_id': site_id_str,
                'site_name': site.site_name,
                'employee_id': str(employee.id),
                'employee_name': employee.full_name,
                'status': 'failed',
                'reason': 'Failed to generate access token'
            })
            continue

        expires_at = datetime.now(timezone.utc) + timedelta(days=30)

        access_request = access_repo.create(
            token=token,
            business_id=g.business_id,
            site_id=site.id,
            employee_id=employee.id,
            processing_month=month,
            created_by_user_id=g.current_user.id,
            expires_at=expires_at,
            is_active=True
        )

        url = _build_access_link_url(token)
        message_body = (
            f"שלום {employee.full_name},\n"
            f"להלן הקישור להעלאת כרטיסי העבודה עבור חודש {month.strftime('%m/%Y')}:\n"
            f"{url}"
        )

        try:
            message = client.messages.create(
                from_=from_number,
                body=message_body,
                to=f"whatsapp:{formatted_phone}"
            )
            logger.info(f"WhatsApp sent to {formatted_phone}: {message.sid}")
            sent_count += 1
            sites_metrics.increment_whatsapp_batch_outcome('sent')
            results.append({
                'site_id': site_id_str,
                'site_name': site.site_name,
                'employee_id': str(employee.id),
                'employee_name': employee.full_name,
                'request_id': str(access_request.id),
                'status': 'sent'
            })
        except Exception as e:
            logger.exception(f"Twilio error for site {site.id}")
            failed_count += 1
            sites_metrics.increment_whatsapp_batch_outcome('failed')
            results.append({
                'site_id': site_id_str,
                'site_name': site.site_name,
                'employee_id': str(employee.id),
                'employee_name': employee.full_name,
                'request_id': str(access_request.id),
                'status': 'failed',
                'reason': str(e)
            })

    return api_response(data={
        'total_requested': len(site_ids),
        'processing_month': processing_month,
        'sent_count': sent_count,
        'failed_count': failed_count,
        'skipped_count': skipped_count,
        'results': results
    }, message="Batch WhatsApp processed")


@sites_bp.route('/<uuid:site_id>/access-link/<uuid:request_id>/revoke', methods=['POST'])
@token_required
@role_required('ADMIN', 'OPERATOR_MANAGER')
def revoke_access_link(site_id, request_id):
    site = repo.get_by_id(site_id)
    if not site or site.business_id != g.business_id:
        return api_response(status_code=404, message="Site not found", error="Not Found")

    access_request = access_repo.get_by_id(request_id)
    if not access_request or access_request.business_id != g.business_id or str(access_request.site_id) != str(site_id):
        return api_response(status_code=404, message="Access link not found", error="Not Found")

    access_repo.revoke(request_id)
    return api_response(message="Access link revoked")
