import calendar
import logging
from datetime import date
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple
from uuid import UUID

from flask import Blueprint, g, request
from openpyxl import load_workbook

from ..auth_utils import token_required, role_required
from ..extensions import db
from ..models.sites import Site
from ..models.work_cards import WorkCard
from ..repositories.employee_repository import EmployeeRepository
from ..repositories.site_repository import SiteRepository
from ..repositories.work_card_day_entry_repository import WorkCardDayEntryRepository
from ..repositories.work_card_repository import WorkCardRepository
from .sites import STATUS_DAY_LABELS, _safe_sheet_name
from .utils import api_response

logger = logging.getLogger(__name__)

site_hours_import_bp = Blueprint('site_hours_import', __name__, url_prefix='/api/sites')

_site_repo = SiteRepository()
_employee_repo = EmployeeRepository()
_work_card_repo = WorkCardRepository()
_day_entry_repo = WorkCardDayEntryRepository()

LABEL_TO_STATUS: Dict[str, str] = {v: k for k, v in STATUS_DAY_LABELS.items()}
_SATURDAY_LABEL = 'שבת'
_TARIFF_LABEL = 'מחיר לשעה'  # side-column label the export writes next to the per-hour tariff
_ALLOWED_STRINGS = set(LABEL_TO_STATUS.keys()) | {_SATURDAY_LABEL}

_CARD_STATUS_ORDER = ['APPROVED', 'NEEDS_REVIEW', 'PENDING', 'REJECTED']


def _best_card(a: WorkCard, b: WorkCard) -> WorkCard:
    rank_a = _CARD_STATUS_ORDER.index(a.review_status) if a.review_status in _CARD_STATUS_ORDER else 99
    rank_b = _CARD_STATUS_ORDER.index(b.review_status) if b.review_status in _CARD_STATUS_ORDER else 99
    if rank_b < rank_a:
        return b
    if rank_a < rank_b:
        return a
    # same rank → prefer most recent
    return a if (a.created_at or 0) >= (b.created_at or 0) else b


def _validate_site_sheet(
    ws,
    site: Site,
    days_in_month: int,
    business_id: UUID,
    employee_by_passport: Optional[Dict[str, Any]] = None,
) -> Tuple[List[Tuple[int, str, Any]], List[Dict[str, Any]]]:
    """Validate a single worksheet against a site.

    Returns ``(valid_cols, errors)`` where ``valid_cols`` is a list of
    ``(col_index, passport, employee)`` for employee columns that resolved to a
    business employee, and ``errors`` is a flat list of error dicts. A non-empty
    ``errors`` list means the sheet must not be applied. Structural problems
    (wrong A1 header, no employee columns, broken day column) short-circuit and
    are returned on their own, since the rest of the sheet cannot be trusted.

    Employees are matched by passport across the whole business (not just the
    site): the export selects employees by where they have hours, so a worker
    with multi-site hours legitimately appears on a sheet that is not their home
    site. Pass ``employee_by_passport`` to reuse a prebuilt passport→employee map
    (the batch import builds it once); otherwise it is looked up per sheet.
    """
    # --- structure: A1 must be "יום בחודש" ---
    if str(ws.cell(row=1, column=1).value or '').strip() != 'יום בחודש':
        return [], [{
            "type": "structure",
            "message": "מבנה הקובץ אינו תואם את הפורמט הנדרש (ציפייה ל'יום בחודש' בתא A1)",
        }]

    # --- collect employee columns from row 1 (B+) ---
    passport_cols: List[Tuple[int, str]] = []
    col = 2
    while True:
        v = ws.cell(row=1, column=col).value
        if v is None or str(v).strip() == '':
            break
        passport_cols.append((col, str(v).strip()))
        col += 1

    # A header-only sheet (a site with no employees) is exported with no passport
    # columns. That is legitimate — return "no valid columns, no errors" so the
    # caller can treat it as an empty no-op rather than a hard failure.
    if not passport_cols:
        return [], []

    employee_count = len(passport_cols)

    # --- validate day structure (rows 3 .. days_in_month+2, col A) ---
    # Only the actual days of the month have day rows; the row immediately after
    # them holds the "סה\"כ" total, which must NOT be validated as a day.
    structure_errors: List[Dict[str, Any]] = []
    for day in range(1, days_in_month + 1):
        cell_val = ws.cell(row=day + 2, column=1).value
        raw = str(cell_val or '').strip()
        # _format_day_label produces "7-שבת" style; extract numeric prefix
        day_num_str = raw.split('-')[0]
        try:
            day_num = int(day_num_str)
        except ValueError:
            structure_errors.append({
                "type": "structure",
                "message": f"שורה {day + 2} (יום {day}): ערך לא תקין בעמודת ימים ({raw!r})",
            })
            continue
        if day_num != day:
            structure_errors.append({
                "type": "structure",
                "message": f"שורה {day + 2}: ציפייה ליום {day}, נמצא {day_num}",
            })

    if structure_errors:
        return [], structure_errors

    # --- validate tariff ---
    # The export writes the per-hour tariff in the summary block below the
    # "סה\"כ" total row, at a position that shifts with the month length
    # (tariff_row = days_in_month + 5). Locate it by its 'מחיר לשעה' label in the
    # side column rather than a fixed row, so 30/31-day months both work.
    tariff_errors: List[Dict[str, Any]] = []
    value_col = employee_count + 2
    label_col = value_col + 1
    excel_tariff: Optional[float] = None
    for r in range(days_in_month + 3, days_in_month + 9):
        if str(ws.cell(row=r, column=label_col).value or '').strip() == _TARIFF_LABEL:
            tariff_cell = ws.cell(row=r, column=value_col).value
            if tariff_cell is not None:
                try:
                    excel_tariff = float(tariff_cell)
                except (ValueError, TypeError):
                    pass
            break

    site_tariff: Optional[float] = float(site.hourly_tariff) if site.hourly_tariff is not None else None

    if site_tariff is not None and excel_tariff is not None:
        if abs(excel_tariff - site_tariff) > 0.01:
            tariff_errors.append({
                "type": "tariff_mismatch",
                "message": (
                    f"התעריף בקובץ ({excel_tariff:.2f} ₪) שונה מהתעריף המוגדר לאתר ({site_tariff:.2f} ₪)"
                ),
            })
    elif site_tariff is None and excel_tariff is not None:
        tariff_errors.append({
            "type": "tariff_mismatch",
            "message": f"הקובץ מכיל תעריף ({excel_tariff:.2f} ₪) אך לאתר לא מוגדר תעריף",
        })

    # --- validate employees (matched by passport across the business) ---
    if employee_by_passport is None:
        passports = [p for _, p in passport_cols]
        employee_by_passport = {
            (e.passport_id or '').strip(): e
            for e in _employee_repo.get_by_passports(passports, business_id)
            if e.passport_id
        }

    employee_errors: List[Dict[str, Any]] = []
    valid_cols: List[Tuple[int, str, Any]] = []  # (col_index, passport, employee)
    for col_idx, passport in passport_cols:
        employee = employee_by_passport.get(passport)
        if not employee:
            employee_errors.append({
                "type": "unknown_employee",
                "message": f"עובד עם דרכון {passport} לא נמצא במערכת",
                "passport": passport,
            })
        else:
            valid_cols.append((col_idx, passport, employee))

    # --- validate cell values (only within the month's day rows) ---
    cell_errors: List[Dict[str, Any]] = []
    for col_idx, passport, _ in valid_cols:
        for day in range(1, days_in_month + 1):
            cell_val = ws.cell(row=day + 2, column=col_idx).value
            if cell_val is None:
                continue
            if isinstance(cell_val, str):
                stripped = cell_val.strip()
                if stripped and stripped not in _ALLOWED_STRINGS:
                    cell_errors.append({
                        "type": "unrecognized_value",
                        "message": f"ערך לא מוכר '{stripped}' לעובד {passport} ביום {day}",
                        "passport": passport,
                        "day": day,
                        "value": stripped,
                    })

    return valid_cols, tariff_errors + employee_errors + cell_errors


def _apply_site_sheet(
    ws,
    site: Site,
    valid_cols: List[Tuple[int, str, Any]],
    month_date: date,
    days_in_month: int,
    user,
    business_id: UUID,
) -> Dict[str, Any]:
    """Apply a validated worksheet to a site's work cards / day entries.

    Returns a per-site summary. Repository calls commit internally, so callers
    must validate every sheet before applying any (the all-or-nothing guarantee
    is enforced at the validate-before-apply boundary, matching the single-site
    import). On an unexpected error the caller should roll back the session.
    """
    existing_cards = _work_card_repo.get_by_site_month(site.id, month_date, business_id)
    cards_by_employee: Dict[UUID, WorkCard] = {}
    for card in existing_cards:
        emp_id = card.employee_id
        if emp_id not in cards_by_employee:
            cards_by_employee[emp_id] = card
        else:
            cards_by_employee[emp_id] = _best_card(cards_by_employee[emp_id], card)

    updated_cards = 0
    updated_entries = 0
    employee_summaries: List[Dict[str, Any]] = []

    for col_idx, passport, employee in valid_cols:
        emp_id = employee.id

        work_card = cards_by_employee.get(emp_id)
        if work_card is None:
            work_card = _work_card_repo.create(
                business_id=business_id,
                site_id=site.id,
                employee_id=emp_id,
                processing_month=month_date,
                source='MANUAL',
                uploaded_by_user_id=user.id,
                review_status='NEEDS_REVIEW',
            )
            cards_by_employee[emp_id] = work_card

        card_id = work_card.id

        # Pre-load existing entries for this card
        existing_entries = _day_entry_repo.get_by_work_card(card_id)
        entries_by_day: Dict[int, Any] = {e.day_of_month: e for e in existing_entries}

        card_entries_changed = 0
        for day in range(1, days_in_month + 1):
            cell_val = ws.cell(row=day + 2, column=col_idx).value

            new_total_hours: Optional[float] = None
            new_day_status: Optional[str] = None

            if isinstance(cell_val, (int, float)):
                new_total_hours = float(cell_val)
            elif isinstance(cell_val, str):
                stripped = cell_val.strip()
                if stripped in LABEL_TO_STATUS:
                    new_day_status = LABEL_TO_STATUS[stripped]
                # empty string or שבת → both None (clear)

            existing_entry = entries_by_day.get(day)
            if existing_entry:
                _day_entry_repo.update_entry(
                    existing_entry.id,
                    user.id,
                    total_hours=new_total_hours,
                    day_status=new_day_status,
                    from_time=None,
                    to_time=None,
                )
                card_entries_changed += 1
            elif new_total_hours is not None or new_day_status is not None:
                _day_entry_repo.create(
                    work_card_id=card_id,
                    day_of_month=day,
                    total_hours=new_total_hours,
                    day_status=new_day_status,
                    source='MANUAL_IMPORT',
                    is_valid=True,
                    updated_by_user_id=user.id,
                )
                card_entries_changed += 1

        updated_cards += 1
        updated_entries += card_entries_changed
        employee_summaries.append({
            "passport": passport,
            "name": employee.full_name,
            "work_card_id": str(work_card.id),
            "entries_changed": card_entries_changed,
        })

    return {
        "updated_cards": updated_cards,
        "updated_entries": updated_entries,
        "employees": employee_summaries,
    }


def _parse_day_cell(cell_val: Any) -> Tuple[Optional[str], Optional[float], Optional[str]]:
    """Interpret one day cell. Returns ``(kind, total_hours, day_status)`` where
    ``kind`` is 'hours', 'status', or None (empty / Saturday label — no work)."""
    if isinstance(cell_val, (int, float)):
        return 'hours', float(cell_val), None
    if isinstance(cell_val, str):
        stripped = cell_val.strip()
        if stripped in LABEL_TO_STATUS:
            return 'status', None, LABEL_TO_STATUS[stripped]
    return None, None, None


def _aggregate_employee_days(
    to_apply: List[Tuple[Site, Any, List[Tuple[int, str, Any]]]],
    days_in_month: int,
) -> Dict[Any, Dict[str, Any]]:
    """Collect every reported day per employee across all sheets.

    Returns ``{employee_id: {'employee': emp, 'days': {day: [(site, hours, status), ...]}}}``.
    Only days with an actual value (hours or status) are recorded; empty cells and
    Saturday labels are ignored. A day landing on more than one site is a conflict.
    """
    emp_days: Dict[Any, Dict[str, Any]] = {}
    for site, ws, valid_cols in to_apply:
        for col_idx, _passport, employee in valid_cols:
            rec = emp_days.setdefault(employee.id, {'employee': employee, 'days': {}})
            for day in range(1, days_in_month + 1):
                kind, total, status = _parse_day_cell(ws.cell(row=day + 2, column=col_idx).value)
                if kind is None:
                    continue
                rec['days'].setdefault(day, []).append((site, total, status))
    return emp_days


def _detect_cross_site_conflicts(emp_days: Dict[Any, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """A conflict is the same employee having a reported value on the same day at
    more than one site. Returns one error dict per conflicting (employee, day)."""
    conflicts: List[Dict[str, Any]] = []
    for rec in emp_days.values():
        employee = rec['employee']
        for day, entries in sorted(rec['days'].items()):
            site_names = {s.site_name for (s, _t, _st) in entries}
            if len({s.id for (s, _t, _st) in entries}) > 1:
                conflicts.append({
                    "type": "hours_conflict",
                    "message": (
                        f"לעובד {employee.full_name} (דרכון {employee.passport_id}) דווחו שעות "
                        f"ביום {day} ביותר מאתר אחד: {', '.join(sorted(site_names))}"
                    ),
                    "passport": employee.passport_id,
                    "day": day,
                })
    return conflicts


def _pick_primary_site_id(day_map: Dict[int, Tuple[Site, Any, Any]], employee) -> Any:
    """For a new card, choose the base site: the employee's home site if they have
    days there, otherwise the site where they have the most days."""
    counts: Dict[Any, int] = {}
    for (site, _t, _st) in day_map.values():
        counts[site.id] = counts.get(site.id, 0) + 1
    if not counts:
        return employee.site_id
    if employee.site_id in counts:
        return employee.site_id
    return max(counts.items(), key=lambda kv: kv[1])[0]


def _apply_consolidated(
    emp_days: Dict[Any, Dict[str, Any]],
    month_date: date,
    days_in_month: int,
    user,
    business_id: UUID,
    auto_approve: bool,
) -> Dict[str, Any]:
    """Apply a conflict-free aggregation. Each employee is consolidated onto ONE
    work card (best existing month card, else a new card on their primary site);
    each day's hours/status are written with ``attributed_site_id`` set to the
    sheet's site (NULL when it equals the card's own site). Sibling cards' day
    entries for the month are cleared so hours are not double-counted, and the
    base card is approved when ``auto_approve`` is set."""
    site_entry_counts: Dict[Any, int] = {}
    site_employees: Dict[Any, set] = {}
    site_meta: Dict[Any, Tuple[str, Any]] = {}
    cards_touched = 0
    total_entries = 0

    for emp_id, rec in emp_days.items():
        employee = rec['employee']
        # Conflict-free: at most one site per day.
        day_map: Dict[int, Tuple[Site, Any, Any]] = {d: e[0] for d, e in rec['days'].items()}

        existing = [
            c for c in _work_card_repo.get_by_employee_month(emp_id, month_date, business_id)
            if c.review_status != 'SPLITTING'
        ]

        # An all-blank column means "this employee has no hours this month" — an
        # authoritative statement. If they already have a card, clear it and
        # approve it below; if they have no card at all, there is nothing to
        # record, so skip (don't create an empty card).
        if not day_map and not existing:
            continue

        if existing:
            base = existing[0]
            for c in existing[1:]:
                base = _best_card(base, c)
            base_site_id = base.site_id
        else:
            base_site_id = _pick_primary_site_id(day_map, employee)
            base = _work_card_repo.create(
                business_id=business_id,
                site_id=base_site_id,
                employee_id=emp_id,
                processing_month=month_date,
                source='MANUAL',
                uploaded_by_user_id=user.id,
                review_status='NEEDS_REVIEW',
            )
        cards_touched += 1

        # Clear sibling cards' month entries so the base card is the single source.
        for c in existing:
            if c.id == base.id:
                continue
            for e in _day_entry_repo.get_by_work_card(c.id):
                if 1 <= e.day_of_month <= days_in_month:
                    _day_entry_repo.delete(e.id)

        existing_entries = {e.day_of_month: e for e in _day_entry_repo.get_by_work_card(base.id)}
        for day in range(1, days_in_month + 1):
            present = day_map.get(day)
            ex = existing_entries.get(day)
            if present is not None:
                site, total, status = present
                attributed = site.id if (base_site_id is None or site.id != base_site_id) else None
                if ex:
                    _day_entry_repo.update_entry(
                        ex.id, user.id,
                        total_hours=total, day_status=status,
                        from_time=None, to_time=None,
                        attributed_site_id=attributed,
                    )
                else:
                    _day_entry_repo.create(
                        work_card_id=base.id, day_of_month=day,
                        total_hours=total, day_status=status,
                        attributed_site_id=attributed,
                        source='MANUAL_IMPORT', is_valid=True,
                        updated_by_user_id=user.id,
                    )
                total_entries += 1
                site_entry_counts[site.id] = site_entry_counts.get(site.id, 0) + 1
                site_employees.setdefault(site.id, set()).add((employee.passport_id, employee.full_name))
                site_meta[site.id] = (site.site_name, site.site_code)
            elif ex:
                # No work reported anywhere this day — clear any stale entry.
                _day_entry_repo.update_entry(
                    ex.id, user.id,
                    total_hours=None, day_status=None,
                    from_time=None, to_time=None,
                    attributed_site_id=None,
                )

        if auto_approve:
            _work_card_repo.update(base.id, review_status='APPROVED')

    sites_summary = [
        {
            "site_name": site_meta[sid][0],
            "site_code": site_meta[sid][1],
            "updated_entries": cnt,
            "employees": [{"passport": p, "name": n} for (p, n) in sorted(site_employees.get(sid, set()))],
        }
        for sid, cnt in site_entry_counts.items()
    ]
    sites_summary.sort(key=lambda s: (s["site_name"] or ''))

    return {
        "updated_cards": cards_touched,
        "updated_entries": total_entries,
        "sites": sites_summary,
        "site_ids_with_entries": set(site_entry_counts.keys()),
    }


def _parse_month_param() -> Tuple[Optional[date], Optional[int]]:
    """Parse the ``month`` query param (``YYYY-MM``). Returns ``(month_date, days_in_month)``
    or ``(None, None)`` if invalid."""
    month_str = (request.args.get('month') or '').strip()
    try:
        year, month = (int(p) for p in month_str.split('-'))
        month_date = date(year, month, 1)
    except (ValueError, AttributeError):
        return None, None
    days_in_month = calendar.monthrange(year, month)[1]
    return month_date, days_in_month


@site_hours_import_bp.route('/<site_id>/hours-import', methods=['POST'])
@token_required
@role_required('ADMIN')
def import_hours_from_excel(site_id: str):
    user = g.current_user

    # --- month param ---
    month_date, days_in_month = _parse_month_param()
    if month_date is None:
        return api_response(status_code=400, message="פרמטר month לא תקין (נדרש YYYY-MM)", error="Bad Request")

    # --- file ---
    if 'file' not in request.files:
        return api_response(status_code=400, message="לא נבחר קובץ", error="Bad Request")
    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith('.xlsx'):
        return api_response(status_code=400, message="יש להעלות קובץ בפורמט XLSX בלבד", error="Bad Request")
    file_bytes = f.read()

    # --- site ---
    try:
        site_uuid = UUID(site_id)
    except ValueError:
        return api_response(status_code=400, message="מזהה אתר לא תקין", error="Bad Request")
    site = db.session.query(Site).filter_by(id=site_uuid, business_id=g.business_id).first()
    if not site:
        return api_response(status_code=404, message="אתר לא נמצא", error="Not Found")

    # --- parse workbook ---
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception:
        return api_response(status_code=400, message="הקובץ אינו קובץ Excel תקין", error="Bad Request")
    ws = wb.active

    # --- validate ---
    valid_cols, errors = _validate_site_sheet(ws, site, days_in_month, g.business_id)
    if errors:
        return api_response(
            status_code=400,
            message="הקובץ מכיל שגיאות ולא יובא",
            error="Bad Request",
            data={"validation_errors": errors},
        )
    if not valid_cols:
        # Valid header but no employee columns — a single-site upload of an empty
        # sheet is almost certainly the wrong file.
        return api_response(status_code=400, message="לא נמצאו עמודות עובדים בקובץ", error="Bad Request")

    # --- apply ---
    try:
        summary = _apply_site_sheet(ws, site, valid_cols, month_date, days_in_month, user, g.business_id)
    except Exception:
        logger.exception("Failed to apply hours import for site %s month %s", site_id, month_date)
        db.session.rollback()
        return api_response(status_code=500, message="שגיאה פנימית בעדכון הנתונים", error="Internal Server Error")

    return api_response(
        status_code=200,
        message=(
            f"ייבוא הושלם בהצלחה: עודכנו {summary['updated_cards']} כרטיסי עבודה "
            f"ו-{summary['updated_entries']} רשומות יום"
        ),
        data=summary,
    )


@site_hours_import_bp.route('/summary/hours-import-batch', methods=['POST'])
@token_required
@role_required('ADMIN')
def import_hours_batch():
    """Import an all-sites hours summary workbook (one sheet per site).

    Each sheet title is matched to a site by reproducing the export naming
    (``_safe_sheet_name`` over the same sorted, active site list). Validation is
    all-or-nothing: every sheet is validated first and nothing is written unless
    all sheets are clean. A sheet matching no site is a hard error.
    """
    user = g.current_user

    # --- month param ---
    month_date, days_in_month = _parse_month_param()
    if month_date is None:
        return api_response(status_code=400, message="פרמטר month לא תקין (נדרש YYYY-MM)", error="Bad Request")

    # --- file ---
    if 'file' not in request.files:
        return api_response(status_code=400, message="לא נבחר קובץ", error="Bad Request")
    f = request.files['file']
    if not f.filename or not f.filename.lower().endswith('.xlsx'):
        return api_response(status_code=400, message="יש להעלות קובץ בפורמט XLSX בלבד", error="Bad Request")
    file_bytes = f.read()

    # --- parse workbook ---
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception:
        return api_response(status_code=400, message="הקובץ אינו קובץ Excel תקין", error="Bad Request")

    # --- build sheet-title -> site map, reproducing the export sheet naming ---
    sites = _site_repo.get_all_for_business(g.business_id)
    sites = [s for s in sites if s.is_active]
    sites = sorted(
        sites,
        key=lambda s: (
            (s.site_name or '').strip().lower(),
            (s.site_code or '').strip().lower(),
            str(s.id),
        )
    )
    used_sheet_names: set = set()
    sheet_name_to_site: Dict[str, Site] = {}
    for s in sites:
        title = _safe_sheet_name(s.site_name, used_sheet_names)
        # Key by the stripped title: Excel strips trailing whitespace from sheet
        # names on save, so the uploaded title may differ from the raw export name.
        sheet_name_to_site[title.strip()] = s

    # Resolve employees by passport across the whole business once (the export
    # places multi-site workers on sheets that are not their home site).
    employee_by_passport: Dict[str, Any] = {
        (e.passport_id or '').strip(): e
        for e in _employee_repo.get_all_for_business(g.business_id)
        if e.passport_id
    }

    # --- validate every sheet first (all-or-nothing) ---
    all_errors: List[Dict[str, Any]] = []
    to_apply: List[Tuple[Site, Any, List[Tuple[int, str, Any]]]] = []  # (site, ws, valid_cols)

    for ws in wb.worksheets:
        title = (ws.title or '').strip()
        site = sheet_name_to_site.get(title)
        if site is None:
            all_errors.append({
                "type": "unmatched_sheet",
                "message": f"הגיליון '{title}' אינו תואם לאף אתר מוגדר",
                "sheet": title,
            })
            continue

        valid_cols, errors = _validate_site_sheet(
            ws, site, days_in_month, g.business_id, employee_by_passport=employee_by_passport
        )
        for err in errors:
            err = dict(err)
            err["site"] = site.site_name
            err["sheet"] = title
            all_errors.append(err)
        # A matched sheet with no employee columns (site with no employees) is a
        # legitimate empty no-op — skip it; it surfaces under skipped_sites.
        if not errors and valid_cols:
            to_apply.append((site, ws, valid_cols))

    if all_errors:
        return api_response(
            status_code=400,
            message="הקובץ מכיל שגיאות ולא יובא",
            error="Bad Request",
            data={"validation_errors": all_errors},
        )

    # --- aggregate per employee across all sheets and detect cross-site conflicts ---
    emp_days = _aggregate_employee_days(to_apply, days_in_month)
    conflicts = _detect_cross_site_conflicts(emp_days)
    if conflicts:
        return api_response(
            status_code=400,
            message="הקובץ מכיל התנגשויות בין אתרים (עובד עם שעות באותו יום ביותר מאתר אחד) ולא יובא",
            error="Bad Request",
            data={"validation_errors": conflicts},
        )

    # --- apply (consolidate each employee onto one card with per-day site) ---
    try:
        result = _apply_consolidated(
            emp_days, month_date, days_in_month, user, g.business_id, auto_approve=True
        )
    except Exception:
        logger.exception("Failed to apply batch hours import for month %s", month_date)
        db.session.rollback()
        return api_response(status_code=500, message="שגיאה פנימית בעדכון הנתונים", error="Internal Server Error")

    used_site_ids = result.pop("site_ids_with_entries")
    skipped_sites = [
        {"site_name": s.site_name, "site_code": s.site_code}
        for s in sites if s.id not in used_site_ids
    ]
    result["skipped_sites"] = skipped_sites

    return api_response(
        status_code=200,
        message=(
            f"ייבוא הושלם בהצלחה: עודכנו {len(result['sites'])} אתרים, "
            f"{result['updated_cards']} כרטיסי עבודה ו-{result['updated_entries']} רשומות יום"
        ),
        data=result,
    )
