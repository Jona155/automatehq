import calendar
import csv
import re
import unicodedata
from copy import copy
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


STATUS_LABELS = {
    'APPROVED': 'מאושר',
    'NEEDS_REVIEW': 'ממתין לסקירה',
    'NEEDS_ASSIGNMENT': 'ממתין לשיוך',
    'REJECTED': 'נדחה',
    'NO_UPLOAD': 'ללא העלאה',
}


class ExportService:
    def __init__(self, hours_matrix_service):
        self.hours_matrix_service = hours_matrix_service

    @staticmethod
    def safe_label(value: str) -> str:
        if not value:
            return ''
        normalized = unicodedata.normalize('NFKC', value)
        cleaned = []
        for ch in normalized:
            cat = unicodedata.category(ch)
            cleaned.append(ch if cat[0] in {'L', 'N'} else '_')
        label = ''.join(cleaned).strip('_')
        return '_'.join(filter(None, label.split('_')))

    @staticmethod
    def safe_sheet_name(value: str, existing: set) -> str:
        base = value or 'Site'
        for ch in ['\\', '/', '*', '[', ']', ':', '?']:
            base = base.replace(ch, ' ')
        base = ' '.join(base.split()).strip() or 'Site'
        base = base[:31]
        candidate, counter = base, 2
        while candidate in existing:
            suffix = f' {counter}'
            candidate = (base[:31 - len(suffix)] + suffix).rstrip()
            counter += 1
        existing.add(candidate)
        return candidate

    @staticmethod
    def resolve_summary_template_path() -> Path:
        base_dir = Path(__file__).resolve().parents[3]
        templates_dir = base_dir / 'excel_extraction_example'
        preferred = templates_dir / 'שעות עבודה לפי אתרים ינואר 26 (1).xlsx'
        if preferred.exists():
            return preferred
        candidates = sorted(templates_dir.glob('*.xlsx'))
        if not candidates:
            raise FileNotFoundError('No Excel template found in backend/excel_extraction_example')
        return candidates[0]

    @staticmethod
    def resolve_salary_template_path(month_date) -> Path:
        root_dir = Path(__file__).resolve().parents[4]
        templates_dir = root_dir / 'employee_sheet_extraction'
        if not templates_dir.exists():
            raise FileNotFoundError('Template folder employee_sheet_extraction was not found')
        month_token = f"{month_date.month:02d}_{month_date.year}"
        month_candidates, generic_candidates = [], []
        for candidate in sorted(templates_dir.glob('*.xlsx')):
            stem = candidate.stem.lower()
            if 'worker_new_hours_template' not in stem:
                continue
            generic_candidates.append(candidate)
            if month_token in stem or f"{month_date.month}_{month_date.year}" in stem:
                month_candidates.append(candidate)
        if month_candidates:
            return month_candidates[-1]
        if generic_candidates:
            return generic_candidates[-1]
        raise FileNotFoundError(f'No salary template found in {templates_dir} for {month_date.strftime("%Y-%m")}')

    @staticmethod
    def extract_salary_day_columns_map(ws, month_date):
        day_columns = {}
        pattern = re.compile(r'^\s*(\d{1,2})\.(\d{1,2})\s*$')
        for col in range(2, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is None:
                continue
            match = pattern.match(str(value))
            if not match:
                continue
            day = int(match.group(1))
            if 1 <= day <= 31:
                day_columns[day] = col
        if not day_columns:
            raise ValueError('Invalid salary template format: missing day columns on header row')
        return day_columns

    @staticmethod
    def _copy_salary_column_template(ws, source_col, target_col):
        ws.column_dimensions[get_column_letter(target_col)].width = ws.column_dimensions[get_column_letter(source_col)].width
        for row in range(1, ws.max_row + 1):
            source_cell = ws.cell(row=row, column=source_col)
            target_cell = ws.cell(row=row, column=target_col)
            target_cell._style = copy(source_cell._style)
            target_cell.number_format = source_cell.number_format
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)
            target_cell.value = source_cell.value

    def _ensure_salary_template_month_columns(self, ws, month_date):
        parsed = self.extract_salary_day_columns_map(ws, month_date)
        day_start_col = min(parsed.values())
        existing_days = max(parsed.keys())
        target_days = calendar.monthrange(month_date.year, month_date.month)[1]
        if existing_days < target_days:
            cols_to_add = target_days - existing_days
            insert_at = day_start_col + existing_days
            ws.insert_cols(insert_at, cols_to_add)
            for offset in range(cols_to_add):
                self._copy_salary_column_template(ws, insert_at - 1, insert_at + offset)
        elif existing_days > target_days:
            ws.delete_cols(day_start_col + target_days, existing_days - target_days)

        day_columns = {}
        for day in range(1, target_days + 1):
            col = day_start_col + day - 1
            ws.cell(row=1, column=col, value=f'{day}.{month_date.month:02d}')
            day_columns[day] = col
        return day_columns

    @staticmethod
    def find_salary_instruction_row(ws):
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value is not None and 'הוראות' in str(cell_value):
                return row
        raise ValueError('Invalid salary template format: could not find instructions row')

    @staticmethod
    def _copy_salary_row_template(ws, source_row, target_row):
        for col in range(1, ws.max_column + 1):
            source_cell = ws.cell(row=source_row, column=col)
            target_cell = ws.cell(row=target_row, column=col)
            target_cell._style = copy(source_cell._style)
            target_cell.number_format = source_cell.number_format
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)
            target_cell.value = source_cell.value
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    def populate_salary_template_sheet(self, ws, employees, matrix, month_date):
        ws.sheet_view.rightToLeft = True
        day_columns = self._ensure_salary_template_month_columns(ws, month_date)
        days_in_month = calendar.monthrange(month_date.year, month_date.month)[1]
        employee_start_row = 2
        instruction_row = self.find_salary_instruction_row(ws)
        base_template_row = max(employee_start_row, instruction_row - 1)

        available_rows = max(0, instruction_row - employee_start_row)
        needed_rows = len(employees)
        if needed_rows > available_rows:
            rows_to_insert = needed_rows - available_rows
            ws.insert_rows(instruction_row, rows_to_insert)
            for row in range(instruction_row, instruction_row + rows_to_insert):
                self._copy_salary_row_template(ws, base_template_row, row)
            instruction_row += rows_to_insert

        for idx, employee in enumerate(employees):
            row_index = employee_start_row + idx
            ws.cell(row=row_index, column=1, value=(employee.passport_id or '').strip() if employee.passport_id else '')
            employee_days = matrix.get(str(employee.id), {})
            for day, col in day_columns.items():
                hours = employee_days.get(day)
                if hours is None:
                    is_saturday = day <= days_in_month and datetime(month_date.year, month_date.month, day).weekday() == 5
                    ws.cell(row=row_index, column=col).value = 'שבת' if is_saturday else None
                else:
                    ws.cell(row=row_index, column=col).value = round(float(hours), 2)

    def export_summary_csv(self, site, employees, matrix, status_map, month):
        days_in_month = calendar.monthrange(month.year, month.month)[1]
        csv_buffer = StringIO()
        writer = csv.writer(csv_buffer)
        writer.writerow(['employee_name', 'employee_id', 'status', *[str(day) for day in range(1, days_in_month + 1)], 'total_hours'])

        for employee in employees:
            employee_id_str = str(employee.id)
            employee_days = matrix.get(employee_id_str, {})
            total_hours = sum(employee_days.values()) if employee_days else 0
            day_values = [f"{employee_days.get(day):.1f}" if employee_days.get(day) is not None else '' for day in range(1, days_in_month + 1)]
            writer.writerow([
                employee.full_name,
                employee_id_str,
                STATUS_LABELS.get(status_map.get(employee_id_str) or 'NO_UPLOAD', status_map.get(employee_id_str) or 'NO_UPLOAD'),
                *day_values,
                f"{total_hours:.1f}" if total_hours > 0 else '',
            ])

        csv_bytes = csv_buffer.getvalue().encode('utf-8-sig')
        output = BytesIO(csv_bytes)
        output.seek(0)
        return output, f"monthly_summary_{self.safe_label(site.site_name) or str(site.id)}_{month.strftime('%Y-%m')}.csv"

    @staticmethod
    def _format_day_label(year: int, month: int, day: int, days_in_month: int):
        if day <= days_in_month and datetime(year, month, day).weekday() == 5:
            return f'{day}-שבת'
        return day

    @staticmethod
    def _day_fallback_value(year: int, month: int, day: int, days_in_month: int):
        if day > days_in_month:
            return None
        if datetime(year, month, day).weekday() == 5:
            return 'שבת'
        return None

    def populate_template_core_sheet(self, ws, employees, matrix, month_date, style_header, style_body, style_total):
        days_in_month = calendar.monthrange(month_date.year, month_date.month)[1]
        employee_count = len(employees)
        last_data_col = max(2, employee_count + 1)
        clear_max_col = max(last_data_col, ws.max_column)
        for row in range(1, 39):
            for col in range(1, clear_max_col + 1):
                ws.cell(row=row, column=col).value = None

        ws.sheet_view.rightToLeft = True
        ws.cell(row=1, column=1, value='יום בחודש')._style = copy(style_header)
        ws.cell(row=2, column=1, value=None)._style = copy(style_body)

        for idx, employee in enumerate(employees, start=2):
            ws.cell(row=1, column=idx, value=(employee.passport_id or '').strip() if employee.passport_id else '')._style = copy(style_header)
            ws.cell(row=2, column=idx, value=employee.full_name or '')._style = copy(style_body)

        for day in range(1, 32):
            row = day + 2
            ws.cell(row=row, column=1, value=self._format_day_label(month_date.year, month_date.month, day, days_in_month))._style = copy(style_body)
            for idx, employee in enumerate(employees, start=2):
                employee_days = matrix.get(str(employee.id), {})
                value = employee_days.get(day)
                if value is None:
                    value = self._day_fallback_value(month_date.year, month_date.month, day, days_in_month)
                ws.cell(row=row, column=idx, value=value)._style = copy(style_body)

        ws.cell(row=34, column=1, value='סה"כ')._style = copy(style_total)
        for col in range(2, last_data_col + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=34, column=col, value=f'=SUM({col_letter}3:{col_letter}33)')._style = copy(style_total)
