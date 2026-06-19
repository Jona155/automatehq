"""
Missing work-cards reporting service.

Single source of truth for "which employees are still missing work cards this
month", pivotable by field manager or by site, plus the XLSX report used for
WhatsApp/email delivery to field managers.

Each employee is expected to submit a configurable number of work cards per
month (default 2 — one mid/end-of-month, one at the start of the following
month). Coverage status per employee:
    NONE     -> 0 cards uploaded
    PARTIAL  -> 1..expected-1 cards (e.g. "only the first card arrived")
    COMPLETE -> >= expected cards
"""
from __future__ import annotations

import calendar
import re
from datetime import date
from io import BytesIO
from typing import Any, Dict, List, Optional
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from sqlalchemy import func

from ..extensions import db
from ..models.business import Business
from ..models.sites import Site, Employee
from ..models.users import User
from ..models.work_cards import WorkCard

IL_COUNTRY_CODE = '972'
DEFAULT_EXPECTED = 2

STATUS_NONE = 'NONE'
STATUS_PARTIAL = 'PARTIAL'
STATUS_COMPLETE = 'COMPLETE'

# Hebrew label for the "no field manager assigned" bucket.
NO_MANAGER_LABEL = 'ללא מנהל שטח'


def normalize_phone_to_whatsapp(raw: Optional[str]) -> Optional[str]:
    """Normalize a phone string to a WhatsApp JID (``<e164digits>@s.whatsapp.net``).

    Mirrors the contractor-phone normalization in ``api/sites.py``: assumes
    Israel (+972) for local numbers starting with 0. Returns None for
    blank/invalid input.
    """
    if not raw or not isinstance(raw, str):
        return None
    trimmed = raw.strip()
    if not trimmed:
        return None

    has_plus = trimmed.startswith('+')
    has_double_zero = trimmed.startswith('00')
    digits = re.sub(r'\D', '', trimmed)
    if not digits:
        return None

    if has_plus:
        pass
    elif has_double_zero:
        digits = digits[2:]
    elif digits.startswith('0'):
        digits = IL_COUNTRY_CODE + digits[1:]

    if not (8 <= len(digits) <= 15):
        return None
    return f'{digits}@s.whatsapp.net'


def _classify(cards_count: int, expected: int) -> str:
    if cards_count <= 0:
        return STATUS_NONE
    if cards_count >= expected:
        return STATUS_COMPLETE
    return STATUS_PARTIAL


def _last_day_of_month(month: date) -> date:
    return date(month.year, month.month, calendar.monthrange(month.year, month.month)[1])


def effective_threshold(month: date, expected: int, today: Optional[date] = None) -> int:
    """How many cards an employee must have *right now* to not count as missing.

    The bar rises over the month's lifecycle:
      * While the reporting month is still open (today on/before its last day),
        only a fully-missing employee (0 cards) is a gap — a single card is
        enough, so the threshold is 1.
      * Once the month has ended (the following month, including the typical
        catch-up window up to the 10th and beyond), the full expected count is
        required, so an employee with only the first card is flagged.
    """
    today = today or date.today()
    if today <= _last_day_of_month(month):
        return 1
    return expected


def compute_missing(
    business_id: UUID,
    month: date,
    site_ids: Optional[List[UUID]] = None,
    today: Optional[date] = None,
) -> List[Dict[str, Any]]:
    """Per active-employee coverage rows for a business+month.

    Returns one dict per active employee (regardless of how many cards they
    have), so callers can both report gaps and compute coverage ratios.

    ``status`` reflects the *time-aware* threshold (see ``effective_threshold``):
    during the open month a single card clears the bar, while after the month
    ends the full expected count is required. ``expected`` always carries the
    configured monthly target so the UI/report can show the real goal.
    """
    business = db.session.query(Business).filter(Business.id == business_id).first()
    business_default = (
        business.expected_work_cards_per_month
        if business and business.expected_work_cards_per_month
        else DEFAULT_EXPECTED
    )

    # Cards per employee for the month (assigned cards only).
    cards_sub = (
        db.session.query(
            WorkCard.employee_id.label('employee_id'),
            func.count(func.distinct(WorkCard.id)).label('cards_count'),
            func.min(WorkCard.created_at).label('first_uploaded_at'),
        )
        .filter(
            WorkCard.business_id == business_id,
            WorkCard.processing_month == month,
            WorkCard.employee_id.isnot(None),
        )
        .group_by(WorkCard.employee_id)
        .subquery()
    )

    query = (
        db.session.query(
            Employee.id.label('employee_id'),
            Employee.full_name.label('full_name'),
            Employee.passport_id.label('passport_id'),
            Employee.phone_number.label('phone_number'),
            Site.id.label('site_id'),
            Site.site_name.label('site_name'),
            Site.expected_work_cards_per_month.label('site_expected'),
            User.id.label('field_manager_id'),
            User.full_name.label('manager_name'),
            User.phone_number.label('manager_phone'),
            func.coalesce(cards_sub.c.cards_count, 0).label('cards_count'),
            cards_sub.c.first_uploaded_at.label('first_uploaded_at'),
        )
        .outerjoin(Site, Site.id == Employee.site_id)
        .outerjoin(User, User.id == Site.field_manager_id)
        .outerjoin(cards_sub, cards_sub.c.employee_id == Employee.id)
        .filter(
            Employee.business_id == business_id,
            Employee.is_active.is_(True),
        )
    )
    if site_ids:
        query = query.filter(Employee.site_id.in_(site_ids))

    query = query.order_by(Site.site_name.asc().nullslast(), Employee.full_name.asc())

    rows: List[Dict[str, Any]] = []
    for r in query.all():
        expected = int(r.site_expected) if r.site_expected else int(business_default)
        threshold = effective_threshold(month, expected, today)
        cards_count = int(r.cards_count or 0)
        rows.append({
            'employee_id': str(r.employee_id),
            'full_name': r.full_name,
            'passport_id': r.passport_id,
            'phone_number': r.phone_number,
            'site_id': str(r.site_id) if r.site_id else None,
            'site_name': r.site_name,
            'field_manager_id': str(r.field_manager_id) if r.field_manager_id else None,
            'manager_name': r.manager_name,
            'manager_phone': r.manager_phone,
            'cards_count': cards_count,
            'expected': expected,
            'status': _classify(cards_count, threshold),
            'first_uploaded_at': r.first_uploaded_at.isoformat() if r.first_uploaded_at else None,
        })
    return rows


def _is_gap(row: Dict[str, Any]) -> bool:
    return row['status'] != STATUS_COMPLETE


def _bucket_counts(rows: List[Dict[str, Any]]) -> Dict[str, int]:
    none = sum(1 for r in rows if r['status'] == STATUS_NONE)
    partial = sum(1 for r in rows if r['status'] == STATUS_PARTIAL)
    complete = sum(1 for r in rows if r['status'] == STATUS_COMPLETE)
    return {
        'total_employees': len(rows),
        'none': none,
        'partial': partial,
        'complete': complete,
        'missing': none + partial,
    }


def group_by_field_manager(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Group employees under their site's field manager, listing only the gaps.

    Coverage counts (``total_employees``/``complete_count``) span ALL of the
    manager's employees so the UI can show compliance, while ``employees`` lists
    only those with gaps for action. Managers with no gaps are omitted from this
    missing-cards view. Managerless employees fall into a single synthetic
    bucket so they're never silently dropped.
    """
    groups: Dict[Optional[str], Dict[str, Any]] = {}
    for row in rows:
        key = row['field_manager_id']
        if key not in groups:
            groups[key] = {
                'field_manager_id': key,
                'manager_name': row['manager_name'] if key else NO_MANAGER_LABEL,
                'manager_phone': row['manager_phone'] if key else None,
                'all_employees': [],
            }
        groups[key]['all_employees'].append(row)

    result = []
    for grp in groups.values():
        all_emps = grp.pop('all_employees')
        counts = _bucket_counts(all_emps)
        if counts['missing'] == 0:
            continue  # manager has no missing cards -> not shown here
        result.append({
            **grp,
            'total_employees': counts['total_employees'],
            'complete_count': counts['complete'],
            'none_count': counts['none'],
            'partial_count': counts['partial'],
            'missing_count': counts['missing'],
            'employees': [r for r in all_emps if _is_gap(r)],
        })
    # Real managers first (alpha), managerless bucket last.
    result.sort(key=lambda g: (g['field_manager_id'] is None, (g['manager_name'] or '')))
    return result


def group_by_site(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Group by site including zero-coverage sites (employees with no cards).

    Coverage is computed over ALL active employees of the site, so a site that
    nobody uploaded for shows complete=0, missing=<headcount>.
    """
    groups: Dict[Optional[str], Dict[str, Any]] = {}
    for row in rows:
        key = row['site_id']
        if key not in groups:
            groups[key] = {
                'site_id': key,
                'site_name': row['site_name'] if key else NO_MANAGER_LABEL,
                'field_manager_id': row['field_manager_id'],
                'manager_name': row['manager_name'],
                'manager_phone': row['manager_phone'],
                'all_employees': [],
            }
        groups[key]['all_employees'].append(row)

    result = []
    for grp in groups.values():
        all_emps = grp.pop('all_employees')
        counts = _bucket_counts(all_emps)
        result.append({
            **grp,
            'total_employees': counts['total_employees'],
            'complete_count': counts['complete'],
            'none_count': counts['none'],
            'partial_count': counts['partial'],
            'missing_count': counts['missing'],
            # Only employees with gaps are listed for action.
            'employees': [r for r in all_emps if _is_gap(r)],
        })
    # Sites with the most gaps first.
    result.sort(key=lambda g: g['missing_count'], reverse=True)
    return result


# ----- XLSX report -----

_HEADERS = [
    'שם עובד',
    'ת.ז. / דרכון',
    'טלפון',
    'אתר',
    'כרטיסים שהתקבלו',
    'סטטוס',
    'תאריך כרטיס ראשון',
]

_STATUS_HE = {
    STATUS_NONE: 'לא התקבל כרטיס',
    STATUS_PARTIAL: 'כרטיס ראשון בלבד',
    STATUS_COMPLETE: 'הושלם',
}


def _status_he(row: Dict[str, Any]) -> str:
    if row['status'] == STATUS_PARTIAL:
        return f"התקבלו {row['cards_count']} מתוך {row['expected']} (כרטיס ראשון בלבד)"
    return _STATUS_HE.get(row['status'], row['status'])


def generate_missing_cards_xlsx(
    title: str,
    rows: List[Dict[str, Any]],
    month: date,
) -> BytesIO:
    """Build a single-sheet XLSX of employees with missing cards. RTL-friendly."""
    wb = Workbook()
    ws = wb.active
    ws.title = (month.strftime('%Y-%m'))
    ws.sheet_view.rightToLeft = True

    month_label = month.strftime('%Y-%m')
    ws.append([f'כרטיסי עבודה חסרים — {title} — {month_label}'])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_HEADERS))
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='right')

    header_row_idx = 2
    ws.append(_HEADERS)
    header_fill = PatternFill('solid', fgColor='D9E1F2')
    for col in range(1, len(_HEADERS) + 1):
        cell = ws.cell(row=header_row_idx, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='right')

    # Worst gaps (no card at all) first.
    ordered = sorted(rows, key=lambda r: (r['status'] != STATUS_NONE, r.get('site_name') or '', r['full_name'] or ''))
    for row in ordered:
        first_dt = row.get('first_uploaded_at')
        first_label = first_dt.split('T')[0] if first_dt else ''
        ws.append([
            row.get('full_name') or '',
            row.get('passport_id') or '',
            row.get('phone_number') or '',
            row.get('site_name') or '',
            f"{row['cards_count']}/{row['expected']}",
            _status_he(row),
            first_label,
        ])

    widths = [22, 16, 16, 22, 16, 28, 16]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
