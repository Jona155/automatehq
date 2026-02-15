import unittest
import uuid
from datetime import date
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from backend.app.api.sites import (
    _extract_salary_day_columns_map,
    _find_salary_instruction_row,
    _populate_salary_template_sheet,
    _resolve_salary_template_path,
)


class SalaryTemplateExportTests(unittest.TestCase):
    def _load_template_sheet(self):
        template_path = Path(__file__).resolve().parents[2] / 'employee_sheet_extraction' / 'worker_new_hours_template_02_2026.xlsx'
        workbook = load_workbook(template_path)
        return workbook, workbook.worksheets[0]

    def test_resolve_salary_template_path_for_month(self):
        resolved = _resolve_salary_template_path(date(2026, 2, 1))
        self.assertIn('02_2026', resolved.stem)

    def test_extract_day_columns_map(self):
        _, ws = self._load_template_sheet()
        day_columns = _extract_salary_day_columns_map(ws, date(2026, 2, 1))
        self.assertEqual(day_columns[1], 2)
        self.assertEqual(day_columns[7], 8)
        self.assertEqual(day_columns[28], 29)

    def test_populate_sheet_writes_ids_and_hours_and_preserves_saturday(self):
        _, ws = self._load_template_sheet()
        day_columns = _extract_salary_day_columns_map(ws, date(2026, 2, 1))

        employee_a = SimpleNamespace(id=uuid.uuid4(), passport_id='P-001')
        employee_b = SimpleNamespace(id=uuid.uuid4(), passport_id='P-002')
        matrix = {
            str(employee_a.id): {1: 8.0, 2: 7.5},
            str(employee_b.id): {1: 10.0},
        }

        _populate_salary_template_sheet(
            ws=ws,
            employees=[employee_a, employee_b],
            matrix=matrix,
            month_date=date(2026, 2, 1),
        )

        self.assertEqual(ws.cell(row=2, column=1).value, 'P-001')
        self.assertEqual(ws.cell(row=3, column=1).value, 'P-002')
        self.assertEqual(ws.cell(row=2, column=day_columns[1]).value, 8.0)
        self.assertEqual(ws.cell(row=2, column=day_columns[2]).value, 7.5)
        self.assertEqual(ws.cell(row=3, column=day_columns[1]).value, 10.0)
        self.assertIsNone(ws.cell(row=3, column=day_columns[2]).value)
        self.assertEqual(ws.cell(row=2, column=day_columns[7]).value, '\u05e9\u05d1\u05ea')

    def test_populate_sheet_inserts_rows_before_instructions_when_over_capacity(self):
        _, ws = self._load_template_sheet()
        original_instruction_row = _find_salary_instruction_row(ws)

        employees = [
            SimpleNamespace(id=uuid.uuid4(), passport_id=f'P-{index:03d}')
            for index in range(10)
        ]
        matrix = {}

        _populate_salary_template_sheet(
            ws=ws,
            employees=employees,
            matrix=matrix,
            month_date=date(2026, 2, 1),
        )

        new_instruction_row = _find_salary_instruction_row(ws)
        self.assertGreater(new_instruction_row, original_instruction_row)
        self.assertEqual(ws.cell(row=2, column=1).value, 'P-000')
        self.assertEqual(ws.cell(row=11, column=1).value, 'P-009')

    def test_populate_sheet_adjusts_columns_for_longer_month_and_saturdays(self):
        _, ws = self._load_template_sheet()  # February template (28 days)

        employee = SimpleNamespace(id=uuid.uuid4(), passport_id='P-100')
        matrix = {
            str(employee.id): {29: 8.0}
        }

        _populate_salary_template_sheet(
            ws=ws,
            employees=[employee],
            matrix=matrix,
            month_date=date(2026, 1, 1),  # January has 31 days
        )

        # Column count extended from A+28 to A+31
        self.assertEqual(ws.max_column, 32)
        self.assertEqual(ws.cell(row=1, column=30).value, '29.01')
        self.assertEqual(ws.cell(row=1, column=32).value, '31.01')

        # January 2026 Saturdays: 3, 10, 17, 24, 31
        self.assertEqual(ws.cell(row=2, column=4).value, '\u05e9\u05d1\u05ea')   # day 3
        self.assertEqual(ws.cell(row=2, column=32).value, '\u05e9\u05d1\u05ea')  # day 31

        # Numeric hours should override fallback Saturday text
        self.assertEqual(ws.cell(row=2, column=30).value, 8.0)  # day 29


if __name__ == '__main__':
    unittest.main()
