"""End-to-end tests for the field-manager columns in the Site Details Excel
import/export (backend/app/api/site_tariff_import.py).

The field manager is attached to a site by PHONE (the deterministic key). An
unrecognized phone fails the whole import; an empty phone cell unassigns; a file
without the field-manager column leaves assignments untouched; a name mismatch is
a warning only.
"""
import unittest
import uuid
from io import BytesIO

from dotenv import load_dotenv

load_dotenv()

from openpyxl import Workbook, load_workbook

from backend.app import create_app, db
from backend.app.auth_utils import encode_auth_token
from backend.app.models.business import Business
from backend.app.models.sites import Site
from backend.app.models.users import User


def _build_xlsx(headers, rows):
    """Build an in-memory .xlsx with the given header labels and string rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'תעריפים לכל אתר'
    for ci, label in enumerate(headers, start=1):
        ws.cell(row=1, column=ci, value=label)
    for ri, row in enumerate(rows, start=2):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci)
            cell.number_format = '@'
            cell.value = val
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


class SiteFieldManagerImportTests(unittest.TestCase):
    def setUp(self):
        self.app = create_app()
        self.client = self.app.test_client()
        self.ctx = self.app.app_context()
        self.ctx.push()

        suffix = uuid.uuid4().hex[:8]
        self.business = Business(name=f'FM Biz {suffix}', code=f'fmbiz-{suffix}', is_active=True)
        db.session.add(self.business)
        db.session.flush()

        self.admin = User(
            full_name='Admin', email=f'admin_{suffix}@example.com',
            role='ADMIN', business_id=self.business.id, is_active=True,
        )
        # Field manager attached by phone. Stored in local format with leading 0.
        self.manager = User(
            full_name='מנהל שדה ראשי', phone_number='0501112233',
            role='FIELD_MANAGER', business_id=self.business.id, is_active=True,
        )
        db.session.add_all([self.admin, self.manager])
        db.session.flush()

        # site_a: no manager. site_b: already assigned to manager.
        self.site_a = Site(site_name=f'אתר א {suffix}', business_id=self.business.id, hourly_tariff=50)
        self.site_b = Site(
            site_name=f'אתר ב {suffix}', business_id=self.business.id,
            hourly_tariff=60, field_manager_id=self.manager.id,
        )
        db.session.add_all([self.site_a, self.site_b])
        db.session.commit()

        self.site_a_id = self.site_a.id
        self.site_b_id = self.site_b.id
        self.site_a_name = self.site_a.site_name
        self.site_b_name = self.site_b.site_name
        self.manager_id = self.manager.id

        token = encode_auth_token(str(self.admin.id))
        self.headers = {'Authorization': f'Bearer {token}'}

    def tearDown(self):
        try:
            for s in Site.query.filter_by(business_id=self.business.id).all():
                db.session.delete(s)
            for u in User.query.filter_by(business_id=self.business.id).all():
                db.session.delete(u)
            db.session.delete(Business.query.get(self.business.id))
            db.session.commit()
        except Exception:
            db.session.rollback()
        self.ctx.pop()

    def _preview(self, file_bytes):
        return self.client.post(
            '/api/sites/tariff-import/preview',
            data={'file': (BytesIO(file_bytes), 'sites.xlsx')},
            content_type='multipart/form-data',
            headers=self.headers,
        )

    def _apply(self, rows):
        return self.client.post('/api/sites/tariff-import/apply', json={'rows': rows}, headers=self.headers)

    def _row_for(self, rows, site_name):
        return next(r for r in rows if r['site_name_from_file'] == site_name)

    # ── Export ──────────────────────────────────────────────────────────────
    def test_export_includes_field_manager_columns(self):
        resp = self.client.get('/api/sites/tariff-import/export', headers=self.headers)
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.data))
        ws = wb.active
        header = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        self.assertEqual(header[4], 'מנהל שדה')
        self.assertEqual(header[5], 'טלפון מנהל שדה')
        # site_b row should carry the assigned manager's name + phone.
        b_row = next(r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value == self.site_b_name)
        self.assertEqual(ws.cell(row=b_row, column=5).value, 'מנהל שדה ראשי')
        self.assertEqual(ws.cell(row=b_row, column=6).value, '0501112233')

    # ── Assign by phone ─────────────────────────────────────────────────────
    def test_preview_and_apply_assigns_field_manager(self):
        headers = ['שם האתר', 'טלפון מנהל שדה']
        xlsx = _build_xlsx(headers, [[self.site_a_name, '0501112233']])
        resp = self._preview(xlsx)
        self.assertEqual(resp.status_code, 200)
        data = resp.get_json()['data']
        self.assertFalse(data['blocked'])
        row = self._row_for(data['rows'], self.site_a_name)
        self.assertEqual(row['action'], 'update')
        self.assertEqual(row['new_field_manager_id'], str(self.manager_id))

        apply_resp = self._apply(data['rows'])
        self.assertEqual(apply_resp.status_code, 200)
        db.session.expire_all()
        self.assertEqual(Site.query.get(self.site_a_id).field_manager_id, self.manager_id)

    # ── Whole-import-fail on unrecognized phone ─────────────────────────────
    def test_unrecognized_phone_blocks_whole_import(self):
        headers = ['שם האתר', 'תעריף שעתי', 'טלפון מנהל שדה']
        xlsx = _build_xlsx(headers, [
            [self.site_b_name, '999', ''],            # valid tariff change
            [self.site_a_name, '', '0500000000'],     # unknown FM phone
        ])
        resp = self._preview(xlsx)
        data = resp.get_json()['data']
        self.assertTrue(data['blocked'])
        bad = self._row_for(data['rows'], self.site_a_name)
        self.assertEqual(bad['action'], 'error')
        self.assertTrue(bad['field_manager_unresolved'])

        apply_resp = self._apply(data['rows'])
        self.assertEqual(apply_resp.status_code, 400)
        # Nothing applied — site_b tariff must be unchanged.
        db.session.expire_all()
        self.assertEqual(float(Site.query.get(self.site_b_id).hourly_tariff), 60.0)

    # ── Empty phone cell unassigns ──────────────────────────────────────────
    def test_empty_phone_unassigns_field_manager(self):
        headers = ['שם האתר', 'טלפון מנהל שדה']
        xlsx = _build_xlsx(headers, [[self.site_b_name, '']])
        data = self._preview(xlsx).get_json()['data']
        self.assertFalse(data['blocked'])
        row = self._row_for(data['rows'], self.site_b_name)
        self.assertEqual(row['action'], 'update')
        self.assertIsNone(row['new_field_manager_id'])

        self._apply(data['rows'])
        db.session.expire_all()
        self.assertIsNone(Site.query.get(self.site_b_id).field_manager_id)

    # ── Absent column leaves managers untouched ─────────────────────────────
    def test_absent_field_manager_column_leaves_assignment(self):
        headers = ['שם האתר', 'תעריף שעתי']
        xlsx = _build_xlsx(headers, [[self.site_b_name, '77']])
        data = self._preview(xlsx).get_json()['data']
        row = self._row_for(data['rows'], self.site_b_name)
        self.assertFalse(row['field_manager_present'])

        self._apply(data['rows'])
        db.session.expire_all()
        site_b = Site.query.get(self.site_b_id)
        self.assertEqual(float(site_b.hourly_tariff), 77.0)
        self.assertEqual(site_b.field_manager_id, self.manager_id)  # untouched

    # ── Name mismatch is a warning, not an error ────────────────────────────
    def test_name_mismatch_is_warning_only(self):
        headers = ['שם האתר', 'מנהל שדה', 'טלפון מנהל שדה']
        xlsx = _build_xlsx(headers, [[self.site_a_name, 'שם שגוי לגמרי', '0501112233']])
        data = self._preview(xlsx).get_json()['data']
        self.assertFalse(data['blocked'])
        row = self._row_for(data['rows'], self.site_a_name)
        self.assertEqual(row['action'], 'update')
        self.assertTrue(len(row['warnings']) > 0)
        self.assertEqual(row['new_field_manager_id'], str(self.manager_id))


if __name__ == '__main__':
    unittest.main()
