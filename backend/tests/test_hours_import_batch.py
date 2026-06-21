import io
import unittest
import uuid
from datetime import date

from dotenv import load_dotenv

load_dotenv()

from openpyxl import Workbook

from backend.app import create_app, db
from backend.app.api.sites import _safe_sheet_name
from backend.app.auth_utils import encode_auth_token
from backend.app.models.business import Business
from backend.app.models.sites import Site, Employee
from backend.app.models.users import User
from backend.app.models.work_cards import WorkCard, WorkCardDayEntry

MONTH = '2026-05'
MONTH_DATE = date(2026, 5, 1)


def _build_sheet(ws, passport_to_hours):
    """Populate a worksheet in the import format: A1 header, passport columns,
    day rows 3-33, plain numeric hours for day 1."""
    ws.cell(row=1, column=1, value='יום בחודש')
    col = 2
    for passport, day1_hours in passport_to_hours.items():
        ws.cell(row=1, column=col, value=passport)
        # day labels in column A
        for day in range(1, 32):
            ws.cell(row=day + 2, column=1, value=str(day))
        # day 1 gets hours; rest left empty
        if day1_hours is not None:
            ws.cell(row=3, column=col, value=day1_hours)
        col += 1


def _build_export_like_sheet(ws, passports, days_in_month, tariff=None):
    """Mirror the real all-sites export layout: header, day rows (with a Saturday
    label), a 'סה\"כ' total row right after the last day, and the tariff block."""
    ws.cell(row=1, column=1, value='יום בחודש')
    for i, p in enumerate(passports):
        ws.cell(row=1, column=2 + i, value=p)
    for day in range(1, days_in_month + 1):
        label = f'{day}-שבת' if day == 6 else str(day)
        ws.cell(row=day + 2, column=1, value=label)
    # day-1 gets 8 hours for the first employee; day-6 (Saturday) gets the שבת label
    if passports:
        ws.cell(row=3, column=2, value=8)
        ws.cell(row=8, column=2, value='שבת')
    total_row = days_in_month + 3
    ws.cell(row=total_row, column=1, value='סה"כ')
    for i in range(len(passports)):
        ws.cell(row=total_row, column=2 + i, value=8)  # total hours (must NOT be read as a day)
    if tariff is not None and passports:
        value_col = len(passports) + 2
        ws.cell(row=total_row + 2, column=value_col, value=tariff)
        ws.cell(row=total_row + 2, column=value_col + 1, value='מחיר לשעה')
        ws.cell(row=total_row + 3, column=value_col, value=8 * len(passports) * tariff)
        ws.cell(row=total_row + 3, column=value_col + 1, value='מחיר ללא מע"מ')


def _build_day_sheet(ws, passport_days, days_in_month):
    """passport_days: {passport: {day: hours}} — minimal sheet with day labels."""
    ws.cell(row=1, column=1, value='יום בחודש')
    for i, (passport, daymap) in enumerate(passport_days.items()):
        ws.cell(row=1, column=2 + i, value=passport)
        for day in range(1, days_in_month + 1):
            ws.cell(row=day + 2, column=1, value=str(day))
            if day in daymap:
                ws.cell(row=day + 2, column=2 + i, value=daymap[day])


class HoursImportBatchTests(unittest.TestCase):
    def setUp(self):
        self.app = create_app()
        self.ctx = self.app.app_context()
        self.ctx.push()
        self.client = self.app.test_client()

        suffix = uuid.uuid4().hex[:8]
        self.business = Business(name=f'Biz {suffix}', code=f'biz{suffix}', is_active=True)
        db.session.add(self.business)
        db.session.flush()

        self.user = User(
            business_id=self.business.id,
            full_name='Admin Tester',
            email=f'admin_{suffix}@example.com',
            role='ADMIN',
        )
        db.session.add(self.user)
        db.session.flush()

        # Two active sites; short names so sheet title == site name (no truncation)
        self.site_a = Site(business_id=self.business.id, site_name=f'SiteA {suffix}', site_code='A1', is_active=True)
        self.site_b = Site(business_id=self.business.id, site_name=f'SiteB {suffix}', site_code='B1', is_active=True)
        db.session.add_all([self.site_a, self.site_b])
        db.session.flush()

        self.emp_a = Employee(business_id=self.business.id, site_id=self.site_a.id,
                              full_name='Emp A', passport_id=f'PA{suffix}', is_active=True)
        self.emp_b = Employee(business_id=self.business.id, site_id=self.site_b.id,
                              full_name='Emp B', passport_id=f'PB{suffix}', is_active=True)
        db.session.add_all([self.emp_a, self.emp_b])
        db.session.commit()

        token = encode_auth_token(str(self.user.id))
        self.headers = {'Authorization': f'Bearer {token}'}

    def tearDown(self):
        try:
            biz_id = self.business.id
            card_ids = [c.id for c in WorkCard.query.filter_by(business_id=biz_id).all()]
            if card_ids:
                WorkCardDayEntry.query.filter(WorkCardDayEntry.work_card_id.in_(card_ids)).delete(synchronize_session=False)
            WorkCard.query.filter_by(business_id=biz_id).delete(synchronize_session=False)
            Employee.query.filter_by(business_id=biz_id).delete(synchronize_session=False)
            Site.query.filter_by(business_id=biz_id).delete(synchronize_session=False)
            User.query.filter_by(business_id=biz_id).delete(synchronize_session=False)
            Business.query.filter_by(id=biz_id).delete(synchronize_session=False)
            db.session.commit()
        except Exception:
            db.session.rollback()
        finally:
            self.ctx.pop()

    def _sheet_name(self, site):
        # Reproduce export naming over the sorted active site list
        sites = sorted(
            [self.site_a, self.site_b],
            key=lambda s: ((s.site_name or '').lower(), (s.site_code or '').lower(), str(s.id)),
        )
        used = set()
        mapping = {}
        for s in sites:
            mapping[s.id] = _safe_sheet_name(s.site_name, used)
        return mapping[site.id]

    def _post(self, wb):
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return self.client.post(
            f'/api/sites/summary/hours-import-batch?month={MONTH}',
            headers=self.headers,
            data={'file': (buf, 'all_sites.xlsx')},
            content_type='multipart/form-data',
        )

    def test_happy_path_updates_both_sites(self):
        wb = Workbook()
        wb.remove(wb.active)
        ws_a = wb.create_sheet(title=self._sheet_name(self.site_a))
        _build_sheet(ws_a, {self.emp_a.passport_id: 8})
        ws_b = wb.create_sheet(title=self._sheet_name(self.site_b))
        _build_sheet(ws_b, {self.emp_b.passport_id: 7.5})

        resp = self._post(wb)
        self.assertEqual(resp.status_code, 200, resp.get_json())
        data = resp.get_json()['data']
        self.assertEqual(len(data['sites']), 2)
        self.assertEqual(data['updated_cards'], 2)

        # Day-1 entries were created for both employees
        cards = WorkCard.query.filter_by(business_id=self.business.id).all()
        self.assertEqual(len(cards), 2)
        entries = WorkCardDayEntry.query.filter(
            WorkCardDayEntry.work_card_id.in_([c.id for c in cards]),
            WorkCardDayEntry.day_of_month == 1,
        ).all()
        hours = sorted(float(e.total_hours) for e in entries)
        self.assertEqual(hours, [7.5, 8.0])

    def test_multi_site_consolidation_no_conflict(self):
        # emp_a (home = site_a) appears on BOTH site sheets, on DIFFERENT days.
        wb = Workbook(); wb.remove(wb.active)
        ws_a = wb.create_sheet(title=self._sheet_name(self.site_a))
        _build_day_sheet(ws_a, {self.emp_a.passport_id: {1: 8}}, 31)
        ws_b = wb.create_sheet(title=self._sheet_name(self.site_b))
        _build_day_sheet(ws_b, {self.emp_a.passport_id: {2: 7}}, 31)

        resp = self._post(wb)
        self.assertEqual(resp.status_code, 200, resp.get_json())

        # Consolidated onto ONE card, auto-approved
        cards = WorkCard.query.filter_by(business_id=self.business.id, employee_id=self.emp_a.id).all()
        self.assertEqual(len(cards), 1)
        card = cards[0]
        self.assertEqual(card.review_status, 'APPROVED')
        self.assertEqual(card.site_id, self.site_a.id)  # home site is the base

        entries = {e.day_of_month: e for e in WorkCardDayEntry.query.filter_by(work_card_id=card.id).all()}
        self.assertEqual(float(entries[1].total_hours), 8.0)
        self.assertEqual(float(entries[2].total_hours), 7.0)
        # day on the base site inherits (null); day on the other site is attributed
        self.assertIsNone(entries[1].attributed_site_id)
        self.assertEqual(entries[2].attributed_site_id, self.site_b.id)

    def test_blank_employee_with_existing_card_is_approved_and_cleared(self):
        # An existing NEEDS_REVIEW card whose employee appears as an all-blank
        # column must be cleared and approved (the import is authoritative).
        card = WorkCard(business_id=self.business.id, site_id=self.site_a.id, employee_id=self.emp_a.id,
                        processing_month=MONTH_DATE, source='EXTRACTED', review_status='NEEDS_REVIEW')
        db.session.add(card); db.session.flush()
        db.session.add(WorkCardDayEntry(work_card_id=card.id, day_of_month=1, total_hours=10,
                                        source='EXTRACTED', is_valid=True))
        db.session.commit()
        card_id = card.id

        wb = Workbook(); wb.remove(wb.active)
        ws_a = wb.create_sheet(title=self._sheet_name(self.site_a))
        _build_day_sheet(ws_a, {self.emp_a.passport_id: {}}, 31)  # column present, all blank

        resp = self._post(wb)
        self.assertEqual(resp.status_code, 200, resp.get_json())
        refreshed = db.session.get(WorkCard, card_id)
        self.assertEqual(refreshed.review_status, 'APPROVED')
        entry = WorkCardDayEntry.query.filter_by(work_card_id=card_id, day_of_month=1).first()
        self.assertIsNone(entry.total_hours)

    def test_same_day_multi_site_conflict_blocks_all(self):
        # emp_a has hours on the SAME day at two sites -> conflict, nothing imported.
        wb = Workbook(); wb.remove(wb.active)
        ws_a = wb.create_sheet(title=self._sheet_name(self.site_a))
        _build_day_sheet(ws_a, {self.emp_a.passport_id: {1: 8}}, 31)
        ws_b = wb.create_sheet(title=self._sheet_name(self.site_b))
        _build_day_sheet(ws_b, {self.emp_a.passport_id: {1: 5}}, 31)

        resp = self._post(wb)
        self.assertEqual(resp.status_code, 400, resp.get_json())
        errors = resp.get_json()['data']['validation_errors']
        self.assertTrue(any(e['type'] == 'hours_conflict' for e in errors))
        self.assertEqual(WorkCard.query.filter_by(business_id=self.business.id).count(), 0)

    def test_unknown_employee_aborts_all(self):
        wb = Workbook()
        wb.remove(wb.active)
        ws_a = wb.create_sheet(title=self._sheet_name(self.site_a))
        _build_sheet(ws_a, {self.emp_a.passport_id: 8})  # valid
        ws_b = wb.create_sheet(title=self._sheet_name(self.site_b))
        _build_sheet(ws_b, {'NOPE-PASSPORT': 7})  # unknown employee

        resp = self._post(wb)
        self.assertEqual(resp.status_code, 400, resp.get_json())
        errors = resp.get_json()['data']['validation_errors']
        self.assertTrue(any(e['type'] == 'unknown_employee' for e in errors))

        # All-or-nothing: site A must NOT have been written either
        cards = WorkCard.query.filter_by(business_id=self.business.id).all()
        self.assertEqual(len(cards), 0)

    def test_export_like_30day_with_tariff_empty_and_cross_site(self):
        from backend.app.models.sites import Site as SiteModel
        suffix = self.business.code[-6:]
        # Site C: has a tariff and an employee whose HOME site is site_a (cross-site)
        site_c = SiteModel(business_id=self.business.id, site_name=f'SiteC {suffix}', site_code='C1',
                           hourly_tariff=50, is_active=True)
        # Site D: active but has no employees -> exported as a header-only empty sheet
        site_d = SiteModel(business_id=self.business.id, site_name=f'SiteD {suffix}', site_code='D1', is_active=True)
        db.session.add_all([site_c, site_d])
        db.session.commit()

        june_days = 30
        wb = Workbook()
        wb.remove(wb.active)
        # emp_a's home site is site_a, but here it appears on site_c's sheet (multi-site hours)
        ws_c = wb.create_sheet(title=f'SiteC {suffix}')
        _build_export_like_sheet(ws_c, [self.emp_a.passport_id], june_days, tariff=50)
        ws_d = wb.create_sheet(title=f'SiteD {suffix}')
        _build_export_like_sheet(ws_d, [], june_days)  # empty header-only sheet

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        resp = self.client.post(
            '/api/sites/summary/hours-import-batch?month=2026-06',
            headers=self.headers,
            data={'file': (buf, 'all.xlsx')},
            content_type='multipart/form-data',
        )
        self.assertEqual(resp.status_code, 200, resp.get_json())
        data = resp.get_json()['data']
        # site_c imported; site_d skipped (empty); the total/tariff rows were not misread
        imported = {s['site_name'] for s in data['sites']}
        self.assertIn(f'SiteC {suffix}', imported)
        skipped = {s['site_name'] for s in data['skipped_sites']}
        self.assertIn(f'SiteD {suffix}', skipped)

        # The day-1 (=8h) entry exists; no spurious day-31 entry from the total row
        card = WorkCard.query.filter_by(business_id=self.business.id, site_id=site_c.id,
                                        employee_id=self.emp_a.id).first()
        self.assertIsNotNone(card)
        days = {e.day_of_month: float(e.total_hours) for e in
                WorkCardDayEntry.query.filter_by(work_card_id=card.id).all()
                if e.total_hours is not None}
        self.assertEqual(days.get(1), 8.0)
        self.assertNotIn(31, days)

    def test_tariff_mismatch_is_error(self):
        from backend.app.models.sites import Site as SiteModel
        suffix = self.business.code[-6:]
        site_c = SiteModel(business_id=self.business.id, site_name=f'SiteC {suffix}', site_code='C1',
                           hourly_tariff=50, is_active=True)
        emp_c = Employee(business_id=self.business.id, site_id=site_c.id,
                         full_name='Emp C', passport_id=f'PC{suffix}', is_active=True)
        db.session.add_all([site_c, emp_c]); db.session.commit()

        wb = Workbook(); wb.remove(wb.active)
        ws_c = wb.create_sheet(title=f'SiteC {suffix}')
        _build_export_like_sheet(ws_c, [emp_c.passport_id], 30, tariff=999)  # wrong tariff
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        resp = self.client.post(
            '/api/sites/summary/hours-import-batch?month=2026-06',
            headers=self.headers,
            data={'file': (buf, 'all.xlsx')},
            content_type='multipart/form-data',
        )
        self.assertEqual(resp.status_code, 400, resp.get_json())
        errors = resp.get_json()['data']['validation_errors']
        self.assertTrue(any(e['type'] == 'tariff_mismatch' for e in errors))
        self.assertEqual(WorkCard.query.filter_by(business_id=self.business.id).count(), 0)

    def test_unmatched_sheet_is_error(self):
        wb = Workbook()
        wb.remove(wb.active)
        ws_a = wb.create_sheet(title=self._sheet_name(self.site_a))
        _build_sheet(ws_a, {self.emp_a.passport_id: 8})
        ws_x = wb.create_sheet(title='No Such Site 999')
        _build_sheet(ws_x, {self.emp_a.passport_id: 8})

        resp = self._post(wb)
        self.assertEqual(resp.status_code, 400, resp.get_json())
        errors = resp.get_json()['data']['validation_errors']
        self.assertTrue(any(e['type'] == 'unmatched_sheet' for e in errors))
        self.assertEqual(WorkCard.query.filter_by(business_id=self.business.id).count(), 0)


if __name__ == '__main__':
    unittest.main()
